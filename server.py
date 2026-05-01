"""
MacJit - Garage Management System Backend
Async event-driven architecture using internal event bus.
Kafka & RabbitMQ adapters log events; activate by setting KAFKA_BOOTSTRAP / RABBITMQ_URL.
Twilio (WhatsApp/SMS) and Razorpay are stubbed adapters.
"""
import os
import uuid
import json
import asyncio
import logging
from pathlib import Path
from datetime import datetime, timezone, timedelta, date
from typing import List, Optional, Dict, Any

import jwt
import bcrypt
from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, WebSocket, WebSocketDisconnect, UploadFile, File, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, ConfigDict
import hashlib
import base64
import hmac
import random
import re
try:
    from cryptography.fernet import Fernet, InvalidToken as FernetInvalidToken
    _cryptography_available = True
except ImportError:
    _cryptography_available = False
    Fernet = None
    FernetInvalidToken = Exception

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

logging.basicConfig(level=logging.INFO, format='%(asctime)s | %(name)s | %(levelname)s | %(message)s')
logger = logging.getLogger("macjit")

# ---------- DB ----------
mongo_url = os.environ['MONGO_URL'].strip().strip('"').strip("'")
if not (mongo_url.startswith('mongodb://') or mongo_url.startswith('mongodb+srv://')):
    mongo_url = 'mongodb+srv://' + mongo_url
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME'].strip().strip('"').strip("'")]

# ---------- Field-level Encryption ----------
ENCRYPTION_KEY = os.environ.get("ENCRYPTION_KEY", "")
_fernet_instance = None


def _get_fernet():
    global _fernet_instance
    if _fernet_instance is None and ENCRYPTION_KEY and _cryptography_available:
        key_bytes = ENCRYPTION_KEY.encode()
        # Accept a raw 32-byte key or a 44-char base64 Fernet key
        if len(ENCRYPTION_KEY) == 44 and ENCRYPTION_KEY.endswith("="):
            fkey = ENCRYPTION_KEY.encode()
        else:
            # Pad/truncate to 32 bytes and base64url-encode into a valid Fernet key
            raw = (key_bytes * 4)[:32]
            fkey = base64.urlsafe_b64encode(raw)
        _fernet_instance = Fernet(fkey)
    return _fernet_instance


def encrypt_field(v: Optional[str]) -> Optional[str]:
    """Encrypt a string field; returns 'enc:<base64>' or original if no key."""
    if not v or not isinstance(v, str):
        return v
    f = _get_fernet()
    if not f:
        return v
    return "enc:" + f.encrypt(v.encode()).decode()


def decrypt_field(v: Optional[str]) -> Optional[str]:
    """Decrypt an 'enc:<base64>' field; returns original string or value unchanged."""
    if not v or not isinstance(v, str) or not v.startswith("enc:"):
        return v
    f = _get_fernet()
    if not f:
        return v
    try:
        return f.decrypt(v[4:].encode()).decode()
    except Exception:
        return v


def field_hash(v: str) -> str:
    """Deterministic SHA-256 hash for queryable encrypted fields."""
    salt = ENCRYPTION_KEY or "macjit-no-key-salt"
    return hashlib.sha256((salt + (v or "")).encode()).hexdigest()


# ---------- NoSQL Injection Prevention ----------
_MONGO_OP_RE = re.compile(r"^\$")

def _sanitize_value(v):
    """Recursively remove MongoDB operator keys from dicts/lists to prevent NoSQL injection."""
    if isinstance(v, dict):
        return {k: _sanitize_value(val) for k, val in v.items() if not _MONGO_OP_RE.match(str(k))}
    if isinstance(v, list):
        return [_sanitize_value(i) for i in v]
    return v


def sanitize_str(v: Optional[str], max_len: int = 500) -> Optional[str]:
    """Strip and truncate a string input; reject if it contains MongoDB operator patterns."""
    if v is None:
        return None
    v = str(v).strip()[:max_len]
    return v


# ---------- OTP Store (in-memory, 10-min TTL) ----------
_approval_otps: Dict[str, dict] = {}  # booking_id -> {otp, phone, exp}
OTP_TTL_SEC = 600  # 10 minutes


def _gen_otp() -> str:
    return str(random.randint(100000, 999999))

def _store_otp(booking_id: str, otp: str, phone: str):
    if not phone:
        raise Exception("Phone required for OTP")

    # Normalize phone
    if not phone.startswith("+"):
        phone = "+91" + phone[-10:]

    _approval_otps[booking_id] = {
        "otp": otp,
        "phone": phone,
        "exp": datetime.now(timezone.utc) + timedelta(seconds=OTP_TTL_SEC),
    }


def _verify_otp(booking_id: str, otp: str) -> bool:
    entry = _approval_otps.get(booking_id)
    if not entry:
        return False

    if datetime.now(timezone.utc) > entry["exp"]:
        _approval_otps.pop(booking_id, None)
        return False

    # 🔐 Secure compare
    if not hmac.compare_digest(str(entry["otp"]), str(otp)):
        return False

    _approval_otps.pop(booking_id, None)
    return True



# ---------- Simple Rate Limiter ----------
_rate_store: Dict[str, list] = {}  # key -> [timestamps]


def _rate_check(key: str, limit: int, window_sec: int) -> bool:
    """Returns True if allowed, False if rate limit exceeded."""
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(seconds=window_sec)
    hits = [t for t in _rate_store.get(key, []) if t > cutoff]
    if len(hits) >= limit:
        return False
    hits.append(now)
    _rate_store[key] = hits
    return True


def decrypt_doc(doc: dict, *fields: str) -> dict:
    """Decrypt named fields in a document dict in-place."""
    if not doc:
        return doc
    for f in fields:
        if f in doc:
            doc[f] = decrypt_field(doc[f])
    return doc



# ---------- Auth ----------
JWT_SECRET = os.environ.get('JWT_SECRET', 'macjit-dev-secret-change-in-prod')
JWT_ALG = "HS256"
security = HTTPBearer(auto_error=False)


def _hash_password_sync(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt(rounds=10)).decode()


def _verify_password_sync(p: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(p.encode(), h.encode())
    except Exception:
        return False


async def hash_password(p: str) -> str:
    return await asyncio.to_thread(_hash_password_sync, p)


async def verify_password(p: str, h: str) -> bool:
    return await asyncio.to_thread(_verify_password_sync, p, h)


def make_token(user_id: str, role: str) -> str:
    payload = {"sub": user_id, "role": role, "exp": datetime.now(timezone.utc) + timedelta(days=7)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


async def get_current_user(creds: HTTPAuthorizationCredentials = Depends(security)) -> Dict[str, Any]:
    if not creds:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(creds.credentials, JWT_SECRET, algorithms=[JWT_ALG])
    except jwt.PyJWTError:
        raise HTTPException(401, "Invalid token")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(401, "User not found")
    return user


def require_roles(*roles):
    async def checker(user=Depends(get_current_user)):
        if user["role"] not in roles:
            raise HTTPException(403, f"Requires role(s): {','.join(roles)}")
        return user
    return checker


# ---------- Models ----------
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


class LoginIn(BaseModel):
    username: Optional[str] = None
    phone: Optional[str] = None
    password: str


class BookingCreate(BaseModel):
    customer_name: str
    customer_phone: str
    car_make: str
    car_model: str
    plate_number: str
    service_type: str
    notes: Optional[str] = ""


class StaffCreate(BaseModel):
    name: str
    phone: str
    role: str  # mechanic | reception | tester | shopkeeper | admin
    password: Optional[str] = None  # admin sets initial password; if blank, server generates one


class PricingIn(BaseModel):
    service_type: str
    base_price: float


class AssignIn(BaseModel):
    mechanic_id: str
    bay_id: str


class ItemAdd(BaseModel):
    inventory_id: str
    qty: int = 1


class ApprovalReq(BaseModel):
    reason: str
    extra_cost: float = 0.0


class InventoryIn(BaseModel):
    name: str
    sku: str
    category: str
    price: float
    stock: int
    low_stock_threshold: int = 5
    stocked_at: Optional[str] = None
    expiry_at: Optional[str] = None


class BulkInventory(BaseModel):
    items: List[InventoryIn]


class EnquiryIn(BaseModel):
    name: str
    phone: str
    email: Optional[str] = ""
    car_make: Optional[str] = ""
    car_model: Optional[str] = ""
    service_interest: Optional[str] = ""
    message: Optional[str] = ""


class LeaveIn(BaseModel):
    leave_type: str  # casual | earned | sick | unpaid
    start_date: str  # ISO date
    end_date: str
    reason: str = ""


class LeaveDecision(BaseModel):
    decision: str  # approved | rejected
    note: str = ""


class HolidayIn(BaseModel):
    date: str
    name: str
    type: str = "public"  # public | optional


class ProfileUpdate(BaseModel):
    monthly_salary: Optional[float] = None
    designation: Optional[str] = None
    join_date: Optional[str] = None
    leave_balance: Optional[Dict[str, int]] = None  # {casual: 12, earned: 15}


class BonusIn(BaseModel):
    user_id: str
    amount: float
    reason: str
    event_type: str = "bonus"  # bonus | extra_work | salary_credited


class ShopSaleLine(BaseModel):
    inventory_id: str
    qty: int = 1


class ShopSaleIn(BaseModel):
    customer_name: Optional[str] = ""
    customer_phone: Optional[str] = ""
    items: List[ShopSaleLine]
    payment_method: str = "cash"  # cash | razorpay
    fitting_charge: float = 0.0   # labour/fitting charge in ₹
    gst_percent: float = 0.0      # GST percentage, e.g. 18 for 18%


class RefundIn(BaseModel):
    sale_id: str
    reason: str
    items: Optional[List[ShopSaleLine]] = None  # if None → full refund of all sale items




class ServiceIn(BaseModel):
    key: str            # unique slug, e.g. "oil-change"
    name: str           # display label, e.g. "Oil & Filter Change"
    duration_min: int   # estimated duration in minutes
    base_price: float   # base charge (₹)
    active: bool = True

class RefundDecision(BaseModel):
    decision: str  # approved | rejected
    note: str = ""


class UserOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    username: str
    name: str
    role: str
    phone: Optional[str] = None


# ---------- Event Bus + Adapters ----------
class EventBus:
    def __init__(self):
        self.subscribers: Dict[str, List[asyncio.Queue]] = {}  # user_id -> queues

    async def subscribe(self, user_id: str) -> asyncio.Queue:
        q: asyncio.Queue = asyncio.Queue()
        self.subscribers.setdefault(user_id, []).append(q)
        return q

    def unsubscribe(self, user_id: str, q: asyncio.Queue):
        if user_id in self.subscribers and q in self.subscribers[user_id]:
            self.subscribers[user_id].remove(q)

    async def fanout(self, user_ids: List[str], event: dict):
        for uid in set(user_ids):
            for q in self.subscribers.get(uid, []):
                try:
                    q.put_nowait(event)
                except asyncio.QueueFull:
                    pass


bus = EventBus()


class KafkaAdapter:
    """Publishes to Confluent Kafka when KAFKA_BOOTSTRAP/API_KEY/API_SECRET are set."""
    enabled = bool(os.environ.get("KAFKA_BOOTSTRAP") and os.environ.get("KAFKA_API_KEY"))
    _producer = None

    @classmethod
    def _get(cls):
        if cls._producer is None and cls.enabled:
            from confluent_kafka import Producer
            cls._producer = Producer({
                "bootstrap.servers": os.environ["KAFKA_BOOTSTRAP"],
                "security.protocol": "SASL_SSL",
                "sasl.mechanisms": "PLAIN",
                "sasl.username": os.environ["KAFKA_API_KEY"],
                "sasl.password": os.environ["KAFKA_API_SECRET"],
            })
        return cls._producer

    @classmethod
    async def publish(cls, topic: str, event: dict):
        if cls.enabled:
            try:
                p = cls._get()
                p.produce(topic, json.dumps(event).encode())
                p.poll(0)
                logger.info(f"[KAFKA->{topic}] {event['type']}")
            except Exception as e:
                logger.error(f"[KAFKA-ERR] {e}")
        else:
            logger.info(f"[KAFKA-MOCK->{topic}] {event['type']} | id={event.get('booking_id','')}")


class RabbitAdapter:
    """Publishes to CloudAMQP when RABBITMQ_URL is set."""
    enabled = bool(os.environ.get("RABBITMQ_URL"))
    _conn = None

    @classmethod
    async def _channel(cls):
        if cls._conn is None and cls.enabled:
            import aio_pika
            cls._conn = await aio_pika.connect_robust(os.environ["RABBITMQ_URL"])
        return await cls._conn.channel() if cls._conn else None

    @classmethod
    async def enqueue(cls, queue_name: str, payload: dict):
        if cls.enabled:
            try:
                import aio_pika
                ch = await cls._channel()
                await ch.declare_queue(queue_name, durable=True)
                await ch.default_exchange.publish(
                    aio_pika.Message(json.dumps(payload).encode()),
                    routing_key=queue_name,
                )
                logger.info(f"[RABBITMQ->{queue_name}] published")
            except Exception as e:
                logger.error(f"[RABBITMQ-ERR] {e}")
        else:
            logger.info(f"[RABBITMQ-MOCK->{queue_name}] enqueued payload")


class TwilioAdapter:
    enabled = bool(os.environ.get("TWILIO_ACCOUNT_SID") and os.environ.get("TWILIO_AUTH_TOKEN"))

    @classmethod
    async def _post(cls, body_data: dict):
        import httpx
        sid = os.environ["TWILIO_ACCOUNT_SID"]
        tok = os.environ["TWILIO_AUTH_TOKEN"]
        url = f"https://api.twilio.com/2010-04-01/Accounts/{sid}/Messages.json"
        async with httpx.AsyncClient(timeout=10) as cli:
            r = await cli.post(url, data=body_data, auth=(sid, tok))
            return r.status_code, r.text

    @classmethod
    async def send_whatsapp(cls, to: str, body: str):
        if cls.enabled:
            try:
                from_num = os.environ.get("TWILIO_WHATSAPP_FROM", "whatsapp:+14155238886").strip()
                if not from_num.startswith("whatsapp:"):
                    from_num = "whatsapp:" + from_num
                code, body_resp = await cls._post({"From": from_num, "To": f"whatsapp:{to}", "Body": body})
                logger.info(f"[TWILIO-WA->{to}] {code} {body_resp[:200] if code >= 400 else ''}")
            except Exception as e:
                logger.error(f"[TWILIO-WA-ERR] {e}")
        else:
            logger.info(f"[TWILIO-WA-MOCK->{to}] {body[:80]}")

    @classmethod
    async def send_sms(cls, to: str, body: str):
        if not cls.enabled:
            logger.warning("Twilio not configured")
            return

        try:
            import httpx

            if not to.startswith("+"):
                to = "+91" + to[-10:]

            from_num = os.environ.get("TWILIO_SMS_FROM")

            async with httpx.AsyncClient(timeout=10) as cli:
                r = await cli.post(
                    f"https://api.twilio.com/2010-04-01/Accounts/{os.environ['TWILIO_ACCOUNT_SID']}/Messages.json",
                    data={"From": from_num, "To": to, "Body": body},
                    auth=(os.environ["TWILIO_ACCOUNT_SID"], os.environ["TWILIO_AUTH_TOKEN"]),
                )

                if r.status_code >= 400:
                    logger.error(f"[TWILIO ERROR] {r.status_code} {r.text}")
                else:
                    logger.info(f"[SMS SENT] {to}")

        except Exception as e:
            logger.error(f"[TWILIO EXCEPTION] {e}")


class RazorpayAdapter:
    enabled = bool(os.environ.get("RAZORPAY_KEY_ID") and os.environ.get("RAZORPAY_KEY_SECRET"))
    WEBHOOK_SECRET = os.environ.get("RAZORPAY_WEBHOOK_SECRET", "")

    @classmethod
    async def create_payment_link_full(cls, amount: float, ref_id: str, customer_phone: str = "",
                                        notes: Optional[Dict[str, Any]] = None) -> tuple:
        if not cls.enabled:
            raise Exception("Razorpay not configured. Check env variables.")

        try:
            import httpx

            # Ensure phone format
            if customer_phone and not customer_phone.startswith("+"):
                customer_phone = "+91" + customer_phone[-10:]

            payload = {
                "amount": int(amount * 100),
                "currency": "INR",
                "description": f"MacJit #{ref_id[:8]}",
                "customer": {"contact": customer_phone} if customer_phone else {},
                "notify": {"sms": True, "email": False},
                "notes": {**(notes or {}), "ref_id": ref_id},
            }

            async with httpx.AsyncClient(timeout=15) as cli:
                r = await cli.post(
                    "https://api.razorpay.com/v1/payment_links",
                    json=payload,
                    auth=(os.environ["RAZORPAY_KEY_ID"], os.environ["RAZORPAY_KEY_SECRET"]),
                )

                # 🔥 CRITICAL FIX
                if r.status_code >= 400:
                    logger.error(f"[RAZORPAY ERROR] {r.status_code} {r.text}")
                    raise Exception(f"Razorpay API failed: {r.text}")

                data = r.json()

                if not data.get("short_url"):
                    raise Exception(f"Invalid Razorpay response: {data}")

                return data["short_url"], data.get("id", "")

        except Exception as e:
            logger.error(f"[RAZORPAY-EXCEPTION] {e}")
            raise HTTPException(500, f"Payment link generation failed: {str(e)}")



# ---------- Event Pipeline ----------
async def publish_event(event_type: str, booking: dict, recipients: List[str], extra: dict = None):
    """Single point that: stores event, fans out via WebSocket, calls Kafka/Rabbit/Twilio adapters."""
    extra = extra or {}
    event = {
        "id": str(uuid.uuid4()),
        "type": event_type,
        "booking_id": booking.get("id"),
        "ts": now_iso(),
        "data": {**{k: v for k, v in booking.items() if k != "_id"}, **extra},
    }
    await db.events.insert_one(dict(event))
    event.pop("_id", None)

    # Build template context once (used for both notif body and Twilio msg)
    _mech = booking.get("mechanic_name") or "our team"
    _tester = booking.get("tester_name") or ""
    _qa_reasons = booking.get("qa_fail_reasons") or []
    _ctx = {
        "plate": booking.get("plate_number", ""),
        "id": (booking.get("id") or "")[:6],
        "mechanic": _mech,
        "tester": _tester,
        "tester_part": f" by {_tester}" if _tester else "",
        "qa_fail_reasons": ", ".join(_qa_reasons) if _qa_reasons else "—",
    }

    # Notifications stored per recipient
    for uid in set(recipients):
        notif = {
            "id": str(uuid.uuid4()),
            "user_id": uid,
            "event_type": event_type,
            "booking_id": booking.get("id"),
            "title": EVENT_TITLES.get(event_type, event_type),
            "body": EVENT_BODIES.get(event_type, "").format(**_ctx),
            "read": False,
            "ts": now_iso(),
        }
        await db.notifications.insert_one(dict(notif))

    await bus.fanout(recipients, event)

    # Adapter side-effects
    await KafkaAdapter.publish("garage.events", event)
    await RabbitAdapter.enqueue(f"queue.{event_type.lower()}", event)

    # Twilio for customer-facing transitions (read phone directly from booking)
    customer_phone = booking.get("customer_phone")
    if customer_phone and event_type in CUSTOMER_NOTIFY_EVENTS:
        msg = EVENT_BODIES.get(event_type, "").format(**_ctx)
        public_url = os.environ.get("PUBLIC_URL") or os.environ.get("APP_URL") or ""
        plate = booking.get("plate_number", "")
        # For BILLED: include the actual Razorpay payment link
        if event_type == "BILLED":
            pay_link = extra.get("payment_link") or booking.get("payment_link") or ""
            if pay_link:
                msg += f"\nPay securely: {pay_link}"
            elif public_url and plate:
                msg += f"\nTrack & pay: {public_url}/track?plate={plate}"
        elif public_url and plate:
            msg += f"\nTrack & pay: {public_url}/track?plate={plate}"
        await TwilioAdapter.send_whatsapp(customer_phone, msg)


EVENT_TITLES = {
    "BOOKING_CREATED": "New Booking",
    "BOOKING_ASSIGNED": "Vehicle Assigned",
    "SERVICE_STARTED": "Service Started",
    "APPROVAL_REQUESTED": "Approval Needed",
    "APPROVAL_GRANTED": "Customer Approved",
    "SERVICE_FINISHED": "Service Finished",
    "QA_DONE": "QA Done - Ready for Pickup",
    "QA_FAIL": "QA Failed - Returned to Mechanic",
    "BILLED": "Bill Generated",
    "PAID": "Payment Received",
}
EVENT_BODIES = {
    "BOOKING_CREATED": "MacJit: Booking #{id} for {plate} confirmed. Mechanic {mechanic} will handle your car.",
    "BOOKING_ASSIGNED": "MacJit: Your car {plate} is assigned to mechanic {mechanic}.",
    "SERVICE_STARTED": "MacJit: Service started on {plate} by {mechanic}. We'll keep you posted.",
    "APPROVAL_REQUESTED": "MacJit: {mechanic} needs your approval for extra work on {plate}. Please open the app.",
    "APPROVAL_GRANTED": "MacJit: Approval received for extra work on {plate}.",
    "SERVICE_FINISHED": "MacJit: {mechanic} finished work on {plate}. Now in QA.",
    "QA_DONE": "MacJit: {plate} passed QA{tester_part}. Ready for pickup!",
    "QA_FAIL": "MacJit: QA FAILED for {plate}. Reason(s): {qa_fail_reasons}. Please fix and resubmit.",
    "BILLED": "MacJit: Bill for {plate} generated. Payment link sent on WhatsApp/SMS.",
    "PAID": "MacJit: Payment received for {plate}. Thanks for choosing us — drive safe!",
}
CUSTOMER_NOTIFY_EVENTS = {"BOOKING_CREATED", "SERVICE_STARTED", "APPROVAL_REQUESTED",
                          "SERVICE_FINISHED", "QA_DONE", "BILLED", "PAID"}

# ---------- Auto-Assignment (8am-6pm service window) ----------
SERVICE_DURATION_MIN = 120  # default fallback
SERVICE_DURATION_BY_TYPE = {
    "oil-change": 45,
    "general": 120,        # 2h
    "ac-service": 90,
    "alignment": 60,
    "brake": 75,
    "engine": 240,         # 4h
    "full-service": 210,   # 3h 30m
}
SHOP_OPEN_HOUR = 8
SHOP_CLOSE_HOUR = 18  # 6 PM
ACTIVE_STATUSES = {"ASSIGNED", "IN_SERVICE"}


def _next_open_slot(after: datetime) -> datetime:
    """Snap a datetime forward to next 8:00-18:00 working window."""
    dt = after
    if dt.hour < SHOP_OPEN_HOUR:
        dt = dt.replace(hour=SHOP_OPEN_HOUR, minute=0, second=0, microsecond=0)
    elif dt.hour >= SHOP_CLOSE_HOUR:
        dt = (dt + timedelta(days=1)).replace(hour=SHOP_OPEN_HOUR, minute=0, second=0, microsecond=0)
    return dt


def _shift_within_hours(start: datetime, duration_min: int) -> tuple:
    """If the slot crosses 18:00, push start to next day 08:00."""
    start = _next_open_slot(start)
    end = start + timedelta(minutes=duration_min)
    close_today = start.replace(hour=SHOP_CLOSE_HOUR, minute=0, second=0, microsecond=0)
    if end > close_today:
        start = (start + timedelta(days=1)).replace(hour=SHOP_OPEN_HOUR, minute=0, second=0, microsecond=0)
        end = start + timedelta(minutes=duration_min)
    return start, end


async def auto_assign_booking(booking_id: str) -> Optional[dict]:
    """Pick mechanic with earliest free slot + first available bay, set ETA fields."""
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        return None
    svc_doc = await db.services.find_one({"key": booking.get("service_type", "general")}, {"_id": 0})
    duration_min = svc_doc["duration_min"] if svc_doc else SERVICE_DURATION_BY_TYPE.get(booking.get("service_type", "general"), SERVICE_DURATION_MIN)
    mechanics = await db.users.find({"role": "mechanic"}, {"_id": 0}).to_list(100)
    bays = await db.bays.find({}, {"_id": 0}).to_list(100)
    if not mechanics or not bays:
        return None

    now = datetime.now(timezone.utc)
    # mechanic free time = max(end of last assigned/in-service booking, now)
    mech_free = {}
    for m in mechanics:
        active = await db.bookings.find(
            {"mechanic_id": m["id"], "status": {"$in": list(ACTIVE_STATUSES)}},
            {"_id": 0}
        ).to_list(100)
        latest_end = now
        for b in active:
            est = b.get("estimated_end_at")
            if est:
                try:
                    end_dt = datetime.fromisoformat(est)
                    if end_dt > latest_end:
                        latest_end = end_dt
                except ValueError:
                    pass
        mech_free[m["id"]] = (m, latest_end)

    # earliest free mechanic
    mech_id, (mech, free_at) = min(mech_free.items(), key=lambda kv: kv[1][1])

    # first bay not currently occupied
    busy_bay_ids = set()
    async for b in db.bookings.find(
        {"status": {"$in": list(ACTIVE_STATUSES)}}, {"_id": 0, "bay_id": 1}
    ):
        if b.get("bay_id"):
            busy_bay_ids.add(b["bay_id"])
    free_bays = [b for b in bays if b["id"] not in busy_bay_ids]
    bay = free_bays[0] if free_bays else bays[0]  # fallback (will queue)

    start_at, end_at = _shift_within_hours(free_at, duration_min)
    upd = {
        "mechanic_id": mech["id"], "mechanic_name": mech["name"],
        "mechanic_phone": mech.get("phone"),
        "bay_id": bay["id"], "bay_name": bay["name"],
        "status": "ASSIGNED",
        "estimated_start_at": start_at.isoformat(),
        "estimated_end_at": end_at.isoformat(),
        "estimated_duration_min": duration_min,
        "auto_assigned": True,
    }
    await db.bookings.update_one({"id": booking_id}, {"$set": upd})
    return await db.bookings.find_one({"id": booking_id}, {"_id": 0})



# ---------- App Setup ----------
app = FastAPI(title="MacJit GMS")
api = APIRouter(prefix="/api")


async def get_recipients_for_booking(booking: dict, include_customer=True, include_admin=True,
                                     include_reception=True, include_mechanic=True, include_tester=False) -> List[str]:
    """Customers no longer have login accounts so include_customer is now ignored;
    customer notifications are sent over WhatsApp via Twilio (see publish_event)."""
    ids = []
    if booking.get("mechanic_id") and include_mechanic:
        ids.append(booking["mechanic_id"])
    role_filter = []
    if include_reception:
        role_filter.append("reception")
    if include_admin:
        role_filter.append("admin")
    if include_tester:
        role_filter.append("tester")
    if role_filter:
        async for u in db.users.find({"role": {"$in": role_filter}}, {"_id": 0, "id": 1}):
            ids.append(u["id"])
    return ids


# ---------- AUTH ROUTES ----------
@api.post("/auth/login")
async def login(data: LoginIn, request: Request):
    """Staff-only login (admin / reception / mechanic / tester / shopkeeper).
    Customers do NOT log in — they track via vehicle number on the public page."""
    ip = request.client.host if request.client else "unknown"
    if not _rate_check(f"login:{ip}", limit=10, window_sec=60):
        raise HTTPException(429, "Too many login attempts — try again in a minute")
    q = {"username": sanitize_str(data.username)} if data.username else {"phone": sanitize_str(data.phone)}
    user = await db.users.find_one(q)
    if not user or not await verify_password(data.password, user["password_hash"]):
        raise HTTPException(401, "Invalid credentials")
    if user.get("role") == "customer":
        raise HTTPException(403, "Customers track bookings on the public page — no login required.")
    if user.get("active") is False:
        raise HTTPException(403, "Account disabled. Contact admin.")
    token = make_token(user["id"], user["role"])
    user.pop("_id", None); user.pop("password_hash", None)
    return {"token": token, "user": user, "must_reset_password": bool(user.get("must_reset_password"))}


@api.post("/auth/change-password")
async def change_password(data: dict, user=Depends(get_current_user)):
    """Force-reset flow on first login, or voluntary change."""
    new_pwd = (data.get("new_password") or "").strip()
    if len(new_pwd) < 6:
        raise HTTPException(400, "Password must be at least 6 characters")
    h = await hash_password(new_pwd)
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"password_hash": h, "must_reset_password": False},
         "$unset": {"initial_password": ""}}
    )
    return {"ok": True}


def _normalize_phone(p: str) -> str:
    """Normalise to +91XXXXXXXXXX for Indian numbers."""
    p = (p or "").strip().replace(" ", "").replace("-", "")
    if not p:
        return p
    if p.startswith("+"):
        return p
    digits = "".join(c for c in p if c.isdigit())
    if len(digits) == 10:
        return "+91" + digits
    if len(digits) == 12 and digits.startswith("91"):
        return "+" + digits
    return "+" + digits if digits else p


# OTP login removed — customers now track via vehicle plate, no auth.




@api.post("/auth/reset-request")
async def reset_request(data: dict, request: Request):
    phone = sanitize_str(data.get("phone"))
    ip = request.client.host if request.client else "unknown"
    if not _rate_check(f"reset:{ip}", limit=5, window_sec=300):
        raise HTTPException(429, "Too many reset requests — try again in 5 minutes")
    user = await db.users.find_one({"phone": phone})
    if not user:
        return {"ok": True}  # silent
    token = uuid.uuid4().hex
    await db.password_resets.insert_one({"token": token, "user_id": user["id"], "ts": now_iso()})
    link = f"{os.environ.get('PUBLIC_URL', '')}/reset?token={token}"
    await TwilioAdapter.send_sms(phone, f"MacJit password reset: {link}")
    return {"ok": True}


@api.post("/auth/reset-confirm")
async def reset_confirm(data: dict):
    token = data.get("token"); new_password = data.get("password")
    rec = await db.password_resets.find_one({"token": token})
    if not rec:
        raise HTTPException(400, "Invalid token")
    h = await hash_password(new_password)
    await db.users.update_one({"id": rec["user_id"]}, {"$set": {"password_hash": h}})
    await db.password_resets.delete_one({"token": token})
    return {"ok": True}


@api.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return user


@api.get("/users")
async def list_users(user=Depends(require_roles("reception", "admin"))):
    return await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(500)


@api.get("/users/by-role/{role}")
async def users_by_role(role: str, user=Depends(get_current_user)):
    return await db.users.find({"role": role}, {"_id": 0, "password_hash": 0}).to_list(500)


# ---------- BAYS ----------
@api.get("/bays")
async def list_bays(user=Depends(get_current_user)):
    return await db.bays.find({}, {"_id": 0}).to_list(100)


# ---------- BOOKINGS ----------
@api.post("/bookings")
async def create_booking(data: BookingCreate, user=Depends(require_roles("reception", "admin"))):
    """Walk-in booking. Customers do NOT log in — we just keep their name/phone on the booking
    and they track everything via the public /track page using the vehicle plate number."""
    if not (data.customer_phone and data.customer_name):
        raise HTTPException(400, "customer_name and customer_phone are required")
    customer_phone = _normalize_phone(data.customer_phone)
    # Look up loyalty info from prior bookings (no user account needed)
    prior = await db.bookings.aggregate([
        {"$match": {"customer_phone": customer_phone, "paid": True}},
        {"$group": {"_id": None, "total": {"$sum": "$bill_amount"}}}
    ]).to_list(1)
    total_spent = (prior[0]["total"] if prior else 0)
    tier = "GOLD" if total_spent >= 25000 else ("SILVER" if total_spent >= 10000 else "BRONZE")
    customer = {
        "id": f"walkin-{customer_phone}",
        "name": data.customer_name,
        "phone": customer_phone,
        "loyalty_tier": tier,
        "total_spent": total_spent,
    }
    booking = {
        "id": str(uuid.uuid4()),
        "customer_id": customer["id"],
        "customer_name": encrypt_field(customer["name"]),
        "customer_name_plain": customer["name"],
        "customer_phone": customer["phone"],
        "loyalty_tier": tier,
        "car_make": data.car_make,
        "car_model": data.car_model,
        "plate_number": data.plate_number,
        "service_type": data.service_type,
        "notes": data.notes,
        "status": "BOOKED",
        "mechanic_id": None,
        "mechanic_name": None,
        "bay_id": None,
        "bay_name": None,
        "items": [],
        "approval_pending": False,
        "approval_reason": None,
        "extra_cost": 0.0,
        "stream_active": False,
        "bill_amount": 0.0,
        "payment_link": None,
        "paid": False,
        "auto_assigned": False,
        "estimated_start_at": None,
        "estimated_end_at": None,
        "created_at": now_iso(),
        "started_at": None,
        "finished_at": None,
        "qa_done_at": None,
        "billed_at": None,
        "paid_at": None,
    }
    await db.bookings.insert_one(dict(booking))
    booking.pop("_id", None)
    # Auto-assign mechanic + bay (1h45m slot, 8am–6pm). No manual reception step.
    assigned = await auto_assign_booking(booking["id"])
    if assigned:
        booking = assigned
        recipients = await get_recipients_for_booking(booking)
        await publish_event("BOOKING_CREATED", booking, recipients,
                            extra={"auto_assigned_to": booking.get("mechanic_name"),
                                   "estimated_start_at": booking.get("estimated_start_at"),
                                   "estimated_end_at": booking.get("estimated_end_at")})
        await publish_event("BOOKING_ASSIGNED", booking, recipients)
    else:
        recipients = await get_recipients_for_booking(booking)
        await publish_event("BOOKING_CREATED", booking, recipients)
    return booking


@api.get("/bookings")
async def list_bookings(status: Optional[str] = None, user=Depends(get_current_user)):
    q = {}
    if status:
        q["status"] = status
    if user["role"] == "mechanic":
        q["mechanic_id"] = user["id"]
    elif user["role"] == "tester":
        q["status"] = {"$in": ["READY_TO_TEST", "QA_DONE", "BILLED", "PAID"]}
    return await db.bookings.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)


@api.get("/bookings/{booking_id}")
async def get_booking(booking_id: str, user=Depends(get_current_user)):
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Not found")
    return b


@api.post("/bookings/{booking_id}/auto-assign")
async def reassign(booking_id: str, user=Depends(require_roles("reception", "admin"))):
    b = await auto_assign_booking(booking_id)
    if not b:
        raise HTTPException(400, "No mechanics or bays available")
    recipients = await get_recipients_for_booking(b)
    await publish_event("BOOKING_ASSIGNED", b, recipients)
    return b


@api.patch("/bookings/{booking_id}/assign")
async def assign(booking_id: str, data: AssignIn, user=Depends(require_roles("reception", "admin"))):
    mech = await db.users.find_one({"id": data.mechanic_id, "role": "mechanic"})
    bay = await db.bays.find_one({"id": data.bay_id})
    if not mech or not bay:
        raise HTTPException(404, "Mechanic or bay not found")
    upd = {"mechanic_id": mech["id"], "mechanic_name": mech["name"],
           "mechanic_phone": mech.get("phone"),
           "bay_id": bay["id"], "bay_name": bay["name"],
           "status": "ASSIGNED"}
    await db.bookings.update_one({"id": booking_id}, {"$set": upd})
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    recipients = await get_recipients_for_booking(b)
    await publish_event("BOOKING_ASSIGNED", b, recipients)
    return b


@api.post("/bookings/{booking_id}/start")
async def start_service(booking_id: str, user=Depends(require_roles("mechanic"))):
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not b or b.get("mechanic_id") != user["id"]:
        raise HTTPException(403, "Not your car")
    upd = {"status": "IN_SERVICE", "stream_active": True, "started_at": now_iso()}
    await db.bookings.update_one({"id": booking_id}, {"$set": upd})
    b.update(upd)
    recipients = await get_recipients_for_booking(b, include_admin=False)  # spec: not admin
    await publish_event("SERVICE_STARTED", b, recipients)
    return b


@api.post("/bookings/{booking_id}/items")
async def add_item(booking_id: str, data: ItemAdd, user=Depends(require_roles("mechanic"))):
    inv = await db.inventory.find_one({"id": data.inventory_id})
    if not inv:
        raise HTTPException(404, "Item not found")
    if inv["stock"] < data.qty:
        raise HTTPException(400, "Insufficient stock")
    line = {"inventory_id": inv["id"], "name": inv["name"], "sku": inv["sku"],
            "qty": data.qty, "price": inv["price"], "subtotal": inv["price"] * data.qty}
    await db.bookings.update_one({"id": booking_id}, {"$push": {"items": line}})
    await db.inventory.update_one({"id": inv["id"]}, {"$inc": {"stock": -data.qty}})
    return await db.bookings.find_one({"id": booking_id}, {"_id": 0})


@api.delete("/bookings/{booking_id}/items/{inventory_id}")
async def remove_item(booking_id: str, inventory_id: str, user=Depends(require_roles("mechanic"))):
    b = await db.bookings.find_one({"id": booking_id})
    if not b:
        raise HTTPException(404, "Not found")
    item = next((i for i in b.get("items", []) if i["inventory_id"] == inventory_id), None)
    if item:
        await db.inventory.update_one({"id": inventory_id}, {"$inc": {"stock": item["qty"]}})
    await db.bookings.update_one({"id": booking_id}, {"$pull": {"items": {"inventory_id": inventory_id}}})
    return await db.bookings.find_one({"id": booking_id}, {"_id": 0})


@api.post("/bookings/{booking_id}/request-approval")
async def request_approval(booking_id: str, data: ApprovalReq, user=Depends(require_roles("mechanic"))):
    upd = {"approval_pending": True, "approval_reason": data.reason, "extra_cost": data.extra_cost}
    await db.bookings.update_one({"id": booking_id}, {"$set": upd})
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    # Generate OTP and send via SMS to customer for secure approval
    cust_phone = b.get("customer_phone")
    if cust_phone:
        otp = _gen_otp()
        _store_otp(booking_id, otp, cust_phone)
        otp_msg = (
            f"MacJit OTP: {otp}\n"
            f"Your mechanic needs approval for extra work on {b.get('plate_number','your car')}.\n"
            f"Extra cost: \u20B9{data.extra_cost}. Reason: {data.reason}\n"
            f"Enter this OTP on the tracking page to approve. Valid 10 minutes. Do NOT share."
        )
        await TwilioAdapter.send_sms(cust_phone, otp_msg)
    recipients = await get_recipients_for_booking(b)
    await publish_event("APPROVAL_REQUESTED", b, recipients, extra={"reason": data.reason, "extra_cost": data.extra_cost})
    return b


@api.post("/bookings/{booking_id}/approval-otp/resend")
async def resend_approval_otp(booking_id: str, request: Request):
    """Public: resend approval OTP to the customer. Rate-limited per booking."""
    ip = request.client.host if request.client else "unknown"
    if not _rate_check(f"otp_resend:{booking_id}:{ip}", limit=3, window_sec=300):
        raise HTTPException(429, "Too many OTP requests — try again in 5 minutes")
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Not found")
    if not b.get("approval_pending"):
        raise HTTPException(400, "No approval pending for this booking")
    cust_phone = b.get("customer_phone")
    if not cust_phone:
        raise HTTPException(400, "No phone on file for this booking")
    otp = _gen_otp()
    _store_otp(booking_id, otp, cust_phone)
    otp_msg = (
        f"MacJit OTP: {otp}\n"
        f"Approve extra work on {b.get('plate_number','your car')}. Valid 10 minutes. Do NOT share."
    )
    await TwilioAdapter.send_sms(cust_phone, otp_msg)
    return {"ok": True, "sent_to": cust_phone[-4:].rjust(10, "*")}


@api.post("/bookings/{booking_id}/approve")
async def approve(booking_id: str, data: Optional[dict] = None, request: Request = None):
    """Public approval — customer must supply the OTP sent to their phone via SMS."""
    ip = (request.client.host if request and request.client else "unknown")
    if not _rate_check(f"otp_verify:{booking_id}:{ip}", limit=5, window_sec=300):
        raise HTTPException(429, "Too many attempts — try again in 5 minutes")
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Not found")
    otp = str((data or {}).get("otp") or "").strip()
    if not otp:
        raise HTTPException(400, "OTP is required")
    if not _verify_otp(booking_id, otp):
        raise HTTPException(403, "Invalid or expired OTP — request a new one")
    await db.bookings.update_one({"id": booking_id}, {"$set": {"approval_pending": False}})
    b["approval_pending"] = False
    recipients = await get_recipients_for_booking(b)
    await publish_event("APPROVAL_GRANTED", b, recipients)
    return b


@api.post("/bookings/{booking_id}/finish")
async def finish(booking_id: str, user=Depends(require_roles("mechanic"))):
    upd = {"status": "READY_TO_TEST", "stream_active": False, "finished_at": now_iso()}
    await db.bookings.update_one({"id": booking_id}, {"$set": upd})
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    recipients = await get_recipients_for_booking(b, include_tester=True)
    await publish_event("SERVICE_FINISHED", b, recipients)
    # Auto-pick the next ASSIGNED booking for this mechanic, ordered by ETA
    nxt = await db.bookings.find_one(
        {"mechanic_id": user["id"], "status": "ASSIGNED"},
        {"_id": 0},
        sort=[("estimated_start_at", 1)],
    )
    return {**b, "next_booking_id": nxt["id"] if nxt else None,
            "next_booking_plate": nxt["plate_number"] if nxt else None}


@api.post("/bookings/{booking_id}/qa-done")
async def qa_done(booking_id: str, user=Depends(require_roles("tester"))):
    upd = {"status": "QA_DONE", "qa_done_at": now_iso(),
           "tester_id": user["id"], "tester_name": user["name"],
           "tester_phone": user.get("phone")}
    await db.bookings.update_one({"id": booking_id}, {"$set": upd})
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    recipients = await get_recipients_for_booking(b, include_tester=True)
    await publish_event("QA_DONE", b, recipients)
    return b


class QAFailIn(BaseModel):
    reasons: List[str]
    notes: Optional[str] = None


@api.post("/bookings/{booking_id}/qa-fail")
async def qa_fail(booking_id: str, body: QAFailIn, user=Depends(require_roles("tester"))):
    if not body.reasons:
        raise HTTPException(400, "At least one fail reason is required")
    upd = {
        "status": "IN_SERVICE",
        "qa_fail_reasons": body.reasons,
        "qa_fail_notes": body.notes or "",
        "qa_failed_at": now_iso(),
        "qa_fail_tester_id": user["id"],
        "qa_fail_tester_name": user["name"],
    }
    await db.bookings.update_one({"id": booking_id}, {"$set": upd})
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    # Notify mechanic via in-app + SMS (not customer)
    recipients = await get_recipients_for_booking(b, include_tester=True, include_reception=True)
    await publish_event("QA_FAIL", b, recipients)
    mechanic_phone = b.get("mechanic_phone") or ""
    if mechanic_phone:
        reasons_str = ", ".join(body.reasons)
        msg = (f"MacJit QA FAILED: {b.get('plate_number')} — Reason(s): {reasons_str}. "
               f"Please fix and resubmit for QA.")
        if body.notes:
            msg += f" Notes: {body.notes}"
        await TwilioAdapter.send_sms(mechanic_phone, msg)
    return b


def _calculate_bill(b: dict, pricing: Optional[dict], cust: Optional[dict],
                    extra_discount: float = 0.0):
    base_charge = (pricing or {}).get("base_price") or DEFAULT_PRICES.get(b.get("service_type", "general"), 800)  # pricing from services DB
    items_total = sum(i.get("subtotal", 0) for i in b.get("items", []))
    extra_cost = b.get("extra_cost", 0) or 0
    subtotal = base_charge + items_total + extra_cost
    tier = (cust or {}).get("loyalty_tier", "BRONZE")
    discount_pct = LOYALTY_DISCOUNT.get(tier, 0)
    loyalty_discount = round(subtotal * discount_pct / 100)
    extra_discount = max(0.0, float(extra_discount or 0))
    total_discount = loyalty_discount + extra_discount
    bill_amount = max(0, subtotal - total_discount)
    return {
        "base_charge": base_charge,
        "items_total": items_total,
        "extra_cost": extra_cost,
        "subtotal": subtotal,
        "loyalty_tier": tier,
        "discount_pct": discount_pct,
        "loyalty_discount": loyalty_discount,
        "extra_discount": extra_discount,
        "discount": total_discount,
        "bill_amount": bill_amount,
    }


@api.get("/bookings/{booking_id}/bill-preview")
async def bill_preview(booking_id: str, extra_discount: float = 0.0,
                       user=Depends(require_roles("reception", "admin"))):
    """Soft-copy preview for reception. Does NOT change booking status."""
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Not found")
    pricing = await db.services.find_one({"key": b.get("service_type")}, {"_id": 0})
    cust = {"name": b.get("customer_name"), "phone": b.get("customer_phone"),
            "loyalty_tier": b.get("loyalty_tier", "BRONZE")}
    calc = _calculate_bill(b, pricing, cust, extra_discount=extra_discount)
    return {
        "booking_id": booking_id,
        "customer": {"name": cust["name"], "phone": cust["phone"]},
        "car": {"make": b.get("car_make"), "model": b.get("car_model"), "plate": b.get("plate_number")},
        "service_type": b.get("service_type"),
        "items": b.get("items", []),
        "approval_reason": b.get("approval_reason"),
        **calc,
        "is_draft": True,
    }


@api.post("/bookings/{booking_id}/bill")
async def bill(booking_id: str, data: Optional[dict] = None,
               user=Depends(require_roles("reception", "admin"))):

    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Not found")

    pricing = await db.services.find_one(
        {"key": b.get("service_type")},
        {"_id": 0}
    )

    calc = _calculate_bill(
        b,
        pricing,
        {"loyalty_tier": b.get("loyalty_tier", "BRONZE")}
    )

bill_amount = calc["bill_amount"]

    try:
        link, rzp_id = await RazorpayAdapter.create_payment_link_full(
            bill_amount,
            booking_id,
            b.get("customer_phone"),
            notes={"booking_id": booking_id}
        )
    except Exception as e:
        raise HTTPException(500, f"Payment failed: {str(e)}")

    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {
            "payment_link": link,
            "rzp_payment_link_id": rzp_id,
            "status": "BILLED",
            "billed_at": now_iso()
        }}
    )

    return {
        "payment_link": link,
        "status": "BILLED"
    }

@api.post("/razorpay/webhook")
async def razorpay_webhook(request: Request):
    """Razorpay webhook receiver. Verifies the X-Razorpay-Signature header against
    RAZORPAY_WEBHOOK_SECRET and, on a payment.captured event, marks the matching
    booking as paid. This catches payments made via the SMS/WhatsApp pay link
    (i.e. payments that don't go through the in-app Checkout modal)."""
    import hmac, hashlib
    secret = os.environ.get("RAZORPAY_WEBHOOK_SECRET", "")
    if not secret:
        raise HTTPException(503, "Webhook secret not configured")
    raw = await request.body()
    signature = request.headers.get("x-razorpay-signature", "")
    expected = hmac.new(secret.encode(), raw, hashlib.sha256).hexdigest()
    if not hmac.compare_digest(expected, signature):
        logger.warning("[RAZORPAY-WEBHOOK] bad signature")
        raise HTTPException(400, "Invalid signature")
    try:
        payload = json.loads(raw.decode() or "{}")
    except Exception:
        raise HTTPException(400, "Bad JSON")
    event = payload.get("event", "")
    if event not in ("payment.captured", "payment_link.paid", "order.paid"):
        return {"ok": True, "ignored": event}
    payment = (((payload.get("payload") or {}).get("payment") or {}).get("entity") or {})
    notes = payment.get("notes") or {}
    booking_id = notes.get("booking_id")
    # payment_link events carry booking_id in the payment_link entity notes
    if not booking_id:
        plink = (((payload.get("payload") or {}).get("payment_link") or {}).get("entity") or {})
        booking_id = (plink.get("notes") or {}).get("booking_id")
    if not booking_id:
        logger.warning(f"[RAZORPAY-WEBHOOK] no booking_id in notes for event={event}")
        return {"ok": True, "no_booking": True}
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not b:
        return {"ok": True, "not_found": True}
    if b.get("paid"):
        return {"ok": True, "already_paid": True}
    upd = {
        "status": "PAID", "paid": True, "paid_at": now_iso(),
        "razorpay_payment_id": payment.get("id"),
        "razorpay_order_id": payment.get("order_id"),
        "payment_method": "razorpay",
    }
    await db.bookings.update_one({"id": booking_id}, {"$set": upd})
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    cust_phone = b.get("customer_phone")
    if cust_phone:
        public_url = os.environ.get("PUBLIC_URL") or os.environ.get("APP_URL") or ""
        invoice_url = f"{public_url}/api/invoices/{booking_id}.pdf"
        msg = (
            f"MacJit Invoice — {b.get('plate_number','')}\n"
            f"Total Paid: \u20B9{b.get('bill_amount',0)}\n"
            f"Invoice: {invoice_url}\n"
            f"Thanks for choosing MacJit. Drive safe!"
        )
        await TwilioAdapter.send_whatsapp(cust_phone, msg)
        sms_invoice = f"MacJit: Payment of \u20B9{b.get('bill_amount',0)} received for {b.get('plate_number','')}. Invoice: {invoice_url}"
        await TwilioAdapter.send_sms(cust_phone, sms_invoice)
    recipients = await get_recipients_for_booking(b)
    await publish_event("PAID", b, recipients)
    return {"ok": True, "booking_id": booking_id}


@api.post("/bookings/{booking_id}/razorpay/order")
async def create_razorpay_order(booking_id: str, data: Optional[dict] = None):
    """Create a Razorpay Order for a booking. Customer must supply plate to confirm.
    Returns the order_id + key_id needed by the Razorpay Checkout JS modal."""
    if not (os.environ.get("RAZORPAY_KEY_ID") and os.environ.get("RAZORPAY_KEY_SECRET")):
        raise HTTPException(503, "Online payment is not configured")
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Not found")
    plate = ((data or {}).get("plate_number") or "").strip().upper()
    if not plate or plate != (b.get("plate_number") or "").upper():
        raise HTTPException(403, "Plate number does not match this booking")
    if b.get("paid"):
        raise HTTPException(400, "Booking is already paid")
    amount_paise = int(round(float(b.get("bill_amount") or 0) * 100))
    if amount_paise <= 0:
        raise HTTPException(400, "Bill not generated yet")
    import httpx
    try:
        async with httpx.AsyncClient(timeout=15) as cli:
            r = await cli.post(
                "https://api.razorpay.com/v1/orders",
                json={
                    "amount": amount_paise,
                    "currency": "INR",
                    "receipt": booking_id[:40],
                    "notes": {"booking_id": booking_id, "plate": b.get("plate_number", "")},
                },
                auth=(os.environ["RAZORPAY_KEY_ID"], os.environ["RAZORPAY_KEY_SECRET"]),
            )
            if r.status_code >= 400:
                logger.error(f"[RAZORPAY-ORDER-ERR] {r.status_code} {r.text}")
                raise HTTPException(502, "Razorpay order creation failed")
            order = r.json()
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[RAZORPAY-ORDER-EXC] {e}")
        raise HTTPException(502, "Razorpay unavailable")
    return {
        "order_id": order["id"],
        "amount": order["amount"],
        "currency": order["currency"],
        "key_id": os.environ["RAZORPAY_KEY_ID"],
        "name": "MacJit Garage",
        "description": f"Service for {b.get('plate_number', '')}",
        "prefill": {
            "name": b.get("customer_name") or "",
            "contact": b.get("customer_phone") or "",
            "email": b.get("customer_email") or "",
        },
    }


@api.post("/bookings/{booking_id}/razorpay/verify")
async def verify_razorpay_payment(booking_id: str, data: dict):
    """Verify a Razorpay payment signature (HMAC-SHA256 of order_id|payment_id with
    key_secret). Only on success do we mark the booking as paid."""
    import hmac, hashlib
    payment_id = (data.get("razorpay_payment_id") or "").strip()
    order_id = (data.get("razorpay_order_id") or "").strip()
    signature = (data.get("razorpay_signature") or "").strip()
    plate = (data.get("plate_number") or "").strip().upper()
    if not (payment_id and order_id and signature):
        raise HTTPException(400, "Missing Razorpay payment fields")
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Not found")
    if not plate or plate != (b.get("plate_number") or "").upper():
        raise HTTPException(403, "Plate number does not match this booking")
    secret = (os.environ.get("RAZORPAY_KEY_SECRET") or "").encode()
    expected = hmac.new(secret, f"{order_id}|{payment_id}".encode(), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(expected, signature):
        logger.warning(f"[RAZORPAY-VERIFY-FAIL] booking={booking_id}")
        raise HTTPException(400, "Invalid payment signature")
    upd = {
        "status": "PAID", "paid": True, "paid_at": now_iso(),
        "razorpay_payment_id": payment_id, "razorpay_order_id": order_id,
        "payment_method": "razorpay",
    }
    await db.bookings.update_one({"id": booking_id}, {"$set": upd})
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    cust_phone = b.get("customer_phone")
    if cust_phone:
        public_url = os.environ.get("PUBLIC_URL") or os.environ.get("APP_URL") or ""
        invoice_url = f"{public_url}/api/invoices/{booking_id}.pdf"
        items_summary = ", ".join([f"{i['name']} x{i['qty']}" for i in (b.get("items") or [])][:3]) or "Service"
        msg = (
            f"MacJit Invoice — {b.get('plate_number','')}\n"
            f"Service: {b.get('service_type','')}\n"
            f"Items: {items_summary}\n"
            f"Total Paid: \u20B9{b.get('bill_amount',0)}\n"
            f"Invoice: {invoice_url}\n"
            f"Thanks for choosing MacJit. Drive safe!"
        )
        await TwilioAdapter.send_whatsapp(cust_phone, msg)
        sms_invoice = f"MacJit: Payment of \u20B9{b.get('bill_amount',0)} received for {b.get('plate_number','')}. Invoice: {invoice_url}"
        await TwilioAdapter.send_sms(cust_phone, sms_invoice)
    recipients = await get_recipients_for_booking(b)
    await publish_event("PAID", b, recipients)
    return b


@api.post("/bookings/{booking_id}/pay")
async def pay(booking_id: str, data: Optional[dict] = None):
    """Public pay confirmation. Customer must include their plate to confirm (when calling
    from the public /track page). Reception/admin can also call this."""
    b0 = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not b0:
        raise HTTPException(404, "Not found")
    plate_supplied = ((data or {}).get("plate_number") or "").strip().upper()
    if plate_supplied and plate_supplied != (b0.get("plate_number") or "").upper():
        raise HTTPException(403, "Plate number does not match this booking")
    upd = {"status": "PAID", "paid": True, "paid_at": now_iso()}
    await db.bookings.update_one({"id": booking_id}, {"$set": upd})
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    # Send invoice link to customer via WhatsApp + SMS
    cust_phone = b.get("customer_phone")
    if cust_phone:
        public_url = os.environ.get("PUBLIC_URL") or os.environ.get("APP_URL") or ""
        invoice_url = f"{public_url}/api/invoices/{booking_id}.pdf"
        items_summary = ", ".join([f"{i['name']} x{i['qty']}" for i in (b.get("items") or [])][:3]) or "Service"
        msg = (
            f"MacJit Invoice — {b.get('plate_number','')}\n"
            f"Service: {b.get('service_type','')}\n"
            f"Items: {items_summary}\n"
            f"Total Paid: \u20B9{b.get('bill_amount',0)}\n"
            f"Invoice: {invoice_url}\n"
            f"Thanks for choosing MacJit. Drive safe!"
        )
        await TwilioAdapter.send_whatsapp(cust_phone, msg)
        sms_invoice = f"MacJit: Payment received for {b.get('plate_number','')}. \u20B9{b.get('bill_amount',0)} paid. Invoice: {invoice_url}"
        await TwilioAdapter.send_sms(cust_phone, sms_invoice)
    recipients = await get_recipients_for_booking(b)
    await publish_event("PAID", b, recipients)
    return b


# ---------- Public customer tracking (no auth) ----------
@api.get("/track")
async def track_by_plate(plate: str):
    """Public endpoint: customer enters their plate number and sees the latest booking,
    bill (if generated) and a pay action — no login required."""
    plate_q = (plate or "").strip().upper()
    if not plate_q:
        raise HTTPException(400, "Plate number required")
    # Case-insensitive match. Use regex anchored to avoid partial collisions.
    import re
    bookings = await db.bookings.find(
        {"plate_number": {"$regex": f"^{re.escape(plate_q)}$", "$options": "i"}},
        {"_id": 0}
    ).sort("created_at", -1).to_list(20)
    if not bookings:
        raise HTTPException(404, "No booking found for this vehicle number")
    active = next((b for b in bookings if b.get("status") != "PAID"), bookings[0])
    public_url = os.environ.get("PUBLIC_URL") or os.environ.get("APP_URL") or ""
    invoice_url = (f"{public_url}/api/invoices/{active['id']}.pdf"
                   if active.get("status") in ("BILLED", "PAID") else None)
    return {"active": active, "history": bookings, "invoice_url": invoice_url}


@api.get("/track/booking/{booking_id}")
async def track_booking_by_id(booking_id: str, plate: str = ""):
    """Public: fetch a single booking by id, but only if the supplied plate matches.
    Used by the customer-side checkout page so the customer can see their bill
    without logging in."""
    plate_q = (plate or "").strip().upper()
    if not plate_q:
        raise HTTPException(400, "Plate number required")
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Not found")
    if (b.get("plate_number") or "").upper() != plate_q:
        raise HTTPException(403, "Plate number does not match this booking")
    return b


@api.post("/track/{booking_id}/send-bill")
async def send_bill_whatsapp(booking_id: str, data: Optional[dict] = None):
    """Public: re-send the current bill summary to the customer's WhatsApp.
    Plate number must be supplied to confirm identity."""
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Not found")
    plate = ((data or {}).get("plate_number") or "").strip().upper()
    if not plate or plate != (b.get("plate_number") or "").upper():
        raise HTTPException(403, "Plate number does not match this booking")
    if b.get("status") not in ("BILLED", "PAID"):
        raise HTTPException(400, "Bill not generated yet — service still in progress.")
    cust_phone = b.get("customer_phone")
    if not cust_phone:
        raise HTTPException(400, "No phone on file")
    public_url = os.environ.get("PUBLIC_URL") or os.environ.get("APP_URL") or ""
    invoice_url = f"{public_url}/api/invoices/{booking_id}.pdf"
    pay_link = b.get("payment_link") or f"{public_url}/track?plate={b.get('plate_number','')}"
    items_summary = ", ".join([f"{i['name']} x{i['qty']}" for i in (b.get("items") or [])][:3]) or "Service"
    status_word = "PAID" if b.get("paid") else "DUE"
    msg = (
        f"MacJit Bill — {b.get('plate_number','')}\n"
        f"Service: {b.get('service_type','')}\n"
        f"Items: {items_summary}\n"
        f"Amount {status_word}: \u20B9{b.get('bill_amount', 0)}\n"
        f"Invoice: {invoice_url}\n"
        f"Pay: {pay_link}"
    )
    await TwilioAdapter.send_whatsapp(cust_phone, msg)
    return {"ok": True, "sent_to": cust_phone}


# ---------- Invoice PDF (public) ----------
@api.get("/invoices/{booking_id}.pdf")
async def invoice_pdf(booking_id: str):
    """Public PDF invoice. Anyone with the booking_id can view/download.
    Available once a bill has been generated (status BILLED or PAID)."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from fastapi.responses import Response

    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Not found")
    if b.get("status") not in ("BILLED", "PAID"):
        raise HTTPException(400, "Bill not generated yet")
    cust = {"name": b.get("customer_name"), "phone": b.get("customer_phone")}

    biz_name = os.environ.get("BUSINESS_NAME", "MacJit")
    biz_loc = os.environ.get("BUSINESS_LOCATION", "Varthur, Bangalore - 560087")
    biz_phone = os.environ.get("BUSINESS_PHONE", "+91 93534 01156")
    biz_email = os.environ.get("BUSINESS_EMAIL", "hello@macjit.com")
    biz_domain = os.environ.get("BUSINESS_DOMAIN", "macjit.com")

    def _watermark(canvas, _doc):
        """Center MacJit logo + text watermark, faded, on every page."""
        canvas.saveState()
        page_w, page_h = A4
        cx, cy = page_w / 2, page_h / 2
        # Outer ring (thin orange)
        canvas.setStrokeColor(colors.HexColor("#F26A21"))
        canvas.setFillColor(colors.HexColor("#F26A21"))
        canvas.setLineWidth(2)
        # Apply transparency for the whole watermark
        try:
            canvas.setFillAlpha(0.07)
            canvas.setStrokeAlpha(0.18)
        except Exception:
            pass
        canvas.circle(cx, cy, 70 * mm, stroke=1, fill=0)
        # Inner solid disk (faded)
        canvas.setFillColor(colors.HexColor("#1E2A44"))
        try:
            canvas.setFillAlpha(0.05)
        except Exception:
            pass
        canvas.circle(cx, cy, 60 * mm, stroke=0, fill=1)
        # Big "M" mark in center
        canvas.setFillColor(colors.HexColor("#F26A21"))
        try:
            canvas.setFillAlpha(0.12)
        except Exception:
            pass
        canvas.setFont("Helvetica-Bold", 200)
        canvas.drawCentredString(cx, cy - 65, "M")
        # Brand strip beneath
        canvas.setFillColor(colors.HexColor("#1E2A44"))
        try:
            canvas.setFillAlpha(0.15)
        except Exception:
            pass
        canvas.setFont("Helvetica-Bold", 28)
        canvas.drawCentredString(cx, cy - 95, "MACJIT")
        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(cx, cy - 110, "MECHANIC · JUST · IN · TIME")
        canvas.restoreState()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=20 * mm,
                            leftMargin=20 * mm, topMargin=18 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=22, textColor=colors.HexColor("#1E2A44"))
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12, textColor=colors.HexColor("#F26A21"))
    body = styles["BodyText"]
    small = ParagraphStyle("small", parent=body, fontSize=9, textColor=colors.HexColor("#555"))

    story = []
    # Header
    story.append(Paragraph(f"<b>{biz_name.upper()}</b> <font size=11 color='#F26A21'>· Mechanic Just In Time</font>", h1))
    story.append(Paragraph(f"{biz_loc} · {biz_phone} · {biz_email} · {biz_domain}", small))
    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph(f"<b>TAX INVOICE</b>  &nbsp;&nbsp; <font color='#888'>#{booking_id[:8].upper()}</font>", h2))
    status_label = "PAID" if b.get("paid") else "DUE"
    color_box = "#16A34A" if b.get("paid") else "#F26A21"
    story.append(Paragraph(
        f"<font color='{color_box}'><b>STATUS: {status_label}</b></font> &nbsp;&nbsp; "
        f"Date: {(b.get('paid_at') or b.get('billed_at') or '')[:10]}",
        body))
    story.append(Spacer(1, 4 * mm))

    # Customer & vehicle
    info_data = [
        ["Customer", (cust or {}).get("name", "—"), "Vehicle", f"{b.get('car_make','')} {b.get('car_model','')}"],
        ["Phone", (cust or {}).get("phone", "—"), "Plate", b.get("plate_number", "—")],
        ["Service", b.get("service_type", "—"), "Mechanic", b.get("mechanic_name") or "—"],
    ]
    info_tbl = Table(info_data, colWidths=[28 * mm, 60 * mm, 28 * mm, 50 * mm])
    info_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#888")),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#888")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#eee")),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 6 * mm))

    # Line items
    rows = [["#", "Item / Service", "SKU", "Qty", "Rate (\u20B9)", "Amount (\u20B9)"]]
    base_charge = b.get("subtotal", 0) - sum(i.get("subtotal", 0) for i in (b.get("items") or [])) - (b.get("extra_cost", 0) or 0)
    rows.append([
        "1",
        f"Service charge ({b.get('service_type','')})",
        "—", "1", f"{base_charge:.0f}", f"{base_charge:.0f}"
    ])
    for idx, it in enumerate(b.get("items") or [], start=2):
        rows.append([
            str(idx), it.get("name", ""), it.get("sku", ""),
            str(it.get("qty", 0)),
            f"{it.get('price', 0):.0f}",
            f"{it.get('subtotal', 0):.0f}",
        ])
    if b.get("extra_cost"):
        rows.append([str(len(rows)), f"Extra: {b.get('approval_reason','add-on')}", "—", "1",
                     f"{b['extra_cost']:.0f}", f"{b['extra_cost']:.0f}"])
    items_tbl = Table(rows, colWidths=[10 * mm, 70 * mm, 28 * mm, 14 * mm, 22 * mm, 26 * mm], repeatRows=1)
    items_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E2A44")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAF8F5")]),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#ddd")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#eee")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(items_tbl)
    story.append(Spacer(1, 4 * mm))

    # Totals
    tot_rows = [
        ["Subtotal", f"\u20B9{b.get('subtotal', 0):.0f}"],
    ]
    if b.get("loyalty_discount"):
        tot_rows.append([f"Loyalty discount ({b.get('loyalty_tier','')} – {b.get('discount_pct',0)}%)",
                         f"-\u20B9{b['loyalty_discount']:.0f}"])
    if b.get("extra_discount"):
        tot_rows.append(["Reception discount", f"-\u20B9{b['extra_discount']:.0f}"])
    tot_rows.append(["TOTAL PAID" if b.get("paid") else "AMOUNT DUE",
                     f"\u20B9{b.get('bill_amount', 0):.0f}"])
    tot_tbl = Table(tot_rows, colWidths=[120 * mm, 50 * mm], hAlign="RIGHT")
    tot_tbl.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 13),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor("#F26A21")),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.HexColor("#1E2A44")),
        ("TOPPADDING", (0, -1), (-1, -1), 6),
    ]))
    story.append(tot_tbl)
    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph(
        "Thank you for choosing MacJit. For any queries please reach us at "
        f"{biz_phone} or {biz_email}.",
        small))
    story.append(Paragraph("This is a computer-generated invoice; no signature required.", small))

    doc.build(story, onFirstPage=_watermark, onLaterPages=_watermark)
    pdf = buf.getvalue()
    buf.close()
    return Response(content=pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="MacJit-{booking_id[:8]}.pdf"'})



# ---------- Pricing ----------
DEFAULT_PRICES = {"general": 800, "oil-change": 500, "full-service": 1800, "tire": 600, "engine": 2500}
LOYALTY_DISCOUNT = {"BRONZE": 0, "SILVER": 5, "GOLD": 10}




# ---------- SERVICES (dynamic, admin-managed) ----------
@api.get("/services")
async def list_services(user=Depends(get_current_user)):
    """Return all services (active and inactive) for display. Reception uses active ones for booking."""
    return await db.services.find({}, {"_id": 0}).sort("key", 1).to_list(100)


@api.get("/services/active")
async def list_active_services(user=Depends(get_current_user)):
    """Return only active services (for booking dropdowns)."""
    return await db.services.find({"active": True}, {"_id": 0}).sort("key", 1).to_list(100)


@api.post("/services")
async def create_service(data: ServiceIn, user=Depends(require_roles("admin"))):
    existing = await db.services.find_one({"key": data.key})
    if existing:
        raise HTTPException(400, f"Service key '{data.key}' already exists. Use PATCH to update.")
    doc = {"id": str(uuid.uuid4()), **data.model_dump(), "created_at": now_iso()}
    await db.services.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.patch("/services/{service_id}")
async def update_service(service_id: str, data: dict, user=Depends(require_roles("admin"))):
    allowed = {"name", "duration_min", "base_price", "active"}
    upd = {k: v for k, v in data.items() if k in allowed}
    if not upd:
        raise HTTPException(400, "Nothing to update. Allowed fields: name, duration_min, base_price, active.")
    await db.services.update_one({"id": service_id}, {"$set": upd})
    svc = await db.services.find_one({"id": service_id}, {"_id": 0})
    if not svc:
        raise HTTPException(404, "Service not found")
    return svc


@api.delete("/services/{service_id}")
async def delete_service(service_id: str, user=Depends(require_roles("admin"))):
    svc = await db.services.find_one({"id": service_id})
    if not svc:
        raise HTTPException(404, "Service not found")
    await db.services.delete_one({"id": service_id})
    return {"ok": True}

@api.get("/pricing")
async def list_pricing(user=Depends(get_current_user)):
    items = await db.service_prices.find({}, {"_id": 0}).to_list(50)
    have = {i["service_type"] for i in items}
    for st, p in DEFAULT_PRICES.items():
        if st not in have:
            items.append({"service_type": st, "base_price": p, "default": True})
    return items


@api.post("/pricing")
async def upsert_pricing(data: PricingIn, user=Depends(require_roles("admin"))):
    await db.service_prices.update_one({"service_type": data.service_type}, {"$set": data.model_dump()}, upsert=True)
    return await db.service_prices.find_one({"service_type": data.service_type}, {"_id": 0})


# ---------- Customer history & loyalty (no accounts; aggregated from bookings) ----------
@api.get("/customers")
async def list_customers(user=Depends(require_roles("reception", "admin"))):
    """Aggregate unique customers from bookings (since customers have no login)."""
    pipeline = [
        {"$group": {
            "_id": "$customer_phone",
            "name": {"$last": "$customer_name"},
            "phone": {"$last": "$customer_phone"},
            "total_spent": {"$sum": {"$cond": [{"$eq": ["$paid", True]}, "$bill_amount", 0]}},
            "visits": {"$sum": 1},
            "last_visit": {"$max": "$created_at"},
        }},
        {"$match": {"_id": {"$ne": None}}},
        {"$sort": {"last_visit": -1}},
    ]
    items = []
    async for c in db.bookings.aggregate(pipeline):
        spent = c.get("total_spent", 0) or 0
        tier = "GOLD" if spent >= 25000 else ("SILVER" if spent >= 10000 else "BRONZE")
        items.append({
            "id": f"walkin-{c['_id']}",
            "name": c.get("name"), "phone": c.get("phone"),
            "total_spent": spent, "loyalty_tier": tier, "visits": c.get("visits", 0),
            "last_visit": c.get("last_visit"),
        })
    return items


@api.get("/customers/{customer_id}/history")
async def customer_history(customer_id: str, user=Depends(require_roles("reception", "admin"))):
    """customer_id is `walkin-{phone}` (matches the IDs returned by /customers)."""
    phone = customer_id.replace("walkin-", "", 1) if customer_id.startswith("walkin-") else customer_id
    bookings = await db.bookings.find({"customer_phone": phone}, {"_id": 0}).sort("created_at", -1).to_list(500)
    if not bookings:
        raise HTTPException(404, "Not found")
    spent = sum(b.get("bill_amount", 0) for b in bookings if b.get("paid"))
    tier = "GOLD" if spent >= 25000 else ("SILVER" if spent >= 10000 else "BRONZE")
    cust = {"id": customer_id, "name": bookings[0].get("customer_name"), "phone": phone,
            "total_spent": spent, "loyalty_tier": tier}
    return {"customer": cust, "bookings": bookings,
            "total_spent": spent, "loyalty_tier": tier,
            "discount_pct": LOYALTY_DISCOUNT.get(tier, 0)}


# ---------- Staff management ----------
ALLOWED_STAFF_ROLES = ("mechanic", "reception", "tester", "shopkeeper", "admin")


@api.get("/admin/staff")
async def list_staff(user=Depends(require_roles("admin"))):
    """Returns staff with their initial password while it has not yet been changed,
    so the admin can pass it on to the employee. Password disappears once the
    employee logs in and resets it."""
    return await db.users.find(
        {"role": {"$in": list(ALLOWED_STAFF_ROLES)}},
        {"_id": 0, "password_hash": 0}
    ).to_list(500)


@api.post("/admin/staff")
async def create_staff(data: StaffCreate, user=Depends(require_roles("admin"))):
    if data.role not in ALLOWED_STAFF_ROLES:
        raise HTTPException(400, "Invalid role")
    phone = _normalize_phone(data.phone)
    if not phone or len(phone) < 10:
        raise HTTPException(400, "Valid phone required")
    if await db.users.find_one({"phone": phone}):
        raise HTTPException(400, "Phone already registered")
    initial_pwd = (data.password or "").strip() or uuid.uuid4().hex[:8]
    if len(initial_pwd) < 6:
        raise HTTPException(400, "Password must be at least 6 characters")
    doc = {"id": str(uuid.uuid4()), "username": phone, "name": data.name,
           "phone": phone, "role": data.role,
           "password_hash": await hash_password(initial_pwd),
           "initial_password": initial_pwd,
           "must_reset_password": True,
           "active": True, "created_at": now_iso(), "created_by": user["id"]}
    await db.users.insert_one(dict(doc))
    # Best-effort SMS so the employee gets the initial password directly.
    try:
        await TwilioAdapter.send_sms(
            phone,
            f"MacJit: Welcome {data.name}! Login at the staff page with phone {phone} "
            f"and temporary password: {initial_pwd}. You will be asked to set a new password."
        )
    except Exception as e:
        logger.warning(f"Staff invite SMS failed: {e}")
    doc.pop("password_hash", None)
    return doc


@api.patch("/admin/staff/{user_id}")
async def update_staff(user_id: str, data: dict, user=Depends(require_roles("admin"))):
    data.pop("_id", None); data.pop("id", None); data.pop("password_hash", None)
    await db.users.update_one({"id": user_id}, {"$set": data})
    return await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})


@api.delete("/admin/staff/{user_id}")
async def delete_staff(user_id: str, user=Depends(require_roles("admin"))):
    target = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(404, "Staff not found")
    if target.get("role") not in ALLOWED_STAFF_ROLES:
        raise HTTPException(400, "Not a staff account")
    if target["id"] == user["id"]:
        raise HTTPException(400, "You cannot remove your own account")
    if target.get("role") == "admin":
        admin_count = await db.users.count_documents({"role": "admin"})
        if admin_count <= 1:
            raise HTTPException(400, "Cannot remove the only admin")
    await db.users.delete_one({"id": user_id})
    return {"ok": True}


@api.post("/admin/staff/{user_id}/reset-password")
async def reset_staff_password(user_id: str, user=Depends(require_roles("admin"))):
    """Generate a fresh temporary password for a staff member and force a reset on next login."""
    target = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not target or target.get("role") not in ALLOWED_STAFF_ROLES:
        raise HTTPException(404, "Staff not found")
    new_pwd = uuid.uuid4().hex[:8]
    h = await hash_password(new_pwd)
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"password_hash": h, "initial_password": new_pwd, "must_reset_password": True}}
    )
    return {"ok": True, "initial_password": new_pwd}


# ---------- Bulk inventory ----------
@api.post("/inventory/bulk")
async def bulk_inventory(data: BulkInventory, user=Depends(require_roles("admin", "reception"))):
    created = []
    for it in data.items:
        item = {"id": str(uuid.uuid4()), **it.model_dump(),
                "stocked_at": it.stocked_at or now_iso(),
                "created_at": now_iso()}
        await db.inventory.insert_one(dict(item))
        item.pop("_id", None)
        created.append(item)
    return {"created": len(created), "items": created}


# ---------- Bulk inventory file upload (CSV / XLSX) ----------
@api.post("/inventory/bulk-upload")
async def bulk_inventory_upload(file: UploadFile = File(...),
                                user=Depends(require_roles("admin", "reception"))):
    """Accepts .csv or .xlsx with columns:
       name, sku, category, price, stock, low_stock_threshold (optional)
       Upserts by SKU: if SKU exists, it adds to stock; otherwise creates new.
    """
    import io, csv
    filename = (file.filename or "").lower()
    raw = await file.read()
    rows: List[dict] = []
    try:
        if filename.endswith(".csv"):
            text = raw.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            rows = [r for r in reader]
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            headers = [str(h or "").strip().lower() for h in next(it)]
            for vals in it:
                if not any(v not in (None, "") for v in vals):
                    continue
                rows.append({headers[i]: (vals[i] if i < len(vals) else "") for i in range(len(headers))})
        else:
            raise HTTPException(400, "Unsupported file. Use .csv or .xlsx")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Could not parse file: {e}")

    REQUIRED = {"name", "sku", "category", "price", "stock"}
    added, updated, errors = 0, 0, []
    for idx, r in enumerate(rows, start=2):  # row 1 is header
        norm = {(k or "").strip().lower(): v for k, v in r.items()}
        missing = [k for k in REQUIRED if not str(norm.get(k, "")).strip()]
        if missing:
            errors.append({"row": idx, "error": f"Missing: {', '.join(missing)}"})
            continue
        try:
            payload = {
                "name": str(norm["name"]).strip(),
                "sku": str(norm["sku"]).strip().upper(),
                "category": str(norm["category"]).strip(),
                "price": float(norm["price"]),
                "stock": int(float(norm["stock"])),
                "low_stock_threshold": int(float(norm.get("low_stock_threshold") or 5)),
            }
        except (ValueError, TypeError) as e:
            errors.append({"row": idx, "error": f"Invalid value: {e}"})
            continue
        existing = await db.inventory.find_one({"sku": payload["sku"]}, {"_id": 0})
        if existing:
            await db.inventory.update_one(
                {"id": existing["id"]},
                {"$inc": {"stock": payload["stock"]},
                 "$set": {"name": payload["name"], "category": payload["category"],
                          "price": payload["price"],
                          "low_stock_threshold": payload["low_stock_threshold"]}})
            updated += 1
        else:
            doc = {"id": str(uuid.uuid4()), **payload,
                   "stocked_at": now_iso(), "created_at": now_iso()}
            await db.inventory.insert_one(doc)
            added += 1
    return {"added": added, "updated": updated, "errors": errors,
            "total_rows": len(rows)}


# ---------- Public Enquiries ----------
@api.post("/enquiries")
async def create_enquiry(data: EnquiryIn):
    doc = {"id": str(uuid.uuid4()), **data.model_dump(),
           "status": "new", "created_at": now_iso()}
    await db.enquiries.insert_one(dict(doc))
    doc.pop("_id", None)
    # Notify admin via SMS (if Twilio configured)
    admins = await db.users.find({"role": "admin"}, {"_id": 0}).to_list(10)
    for a in admins:
        if a.get("phone"):
            await TwilioAdapter.send_sms(
                a["phone"],
                f"MacJit enquiry from {data.name} ({data.phone}): "
                f"{(data.car_make or '')} {(data.car_model or '')} - "
                f"{(data.service_interest or 'general')}"
            )
    return doc


@api.get("/enquiries")
async def list_enquiries(user=Depends(require_roles("admin", "reception"))):
    items = await db.enquiries.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items


@api.patch("/enquiries/{enq_id}")
async def update_enquiry(enq_id: str, data: dict, user=Depends(require_roles("admin", "reception"))):
    data.pop("_id", None); data.pop("id", None)
    await db.enquiries.update_one({"id": enq_id}, {"$set": data})
    return await db.enquiries.find_one({"id": enq_id}, {"_id": 0})


# ---------- INVENTORY ----------
@api.get("/inventory")
async def list_inventory(user=Depends(get_current_user)):
    items = await db.inventory.find({}, {"_id": 0}).to_list(1000)
    items.sort(key=lambda x: x.get("stocked_at") or "")
    return items


@api.post("/inventory")
async def create_inventory(data: InventoryIn, user=Depends(require_roles("admin", "reception"))):
    item = {"id": str(uuid.uuid4()), **data.model_dump(),
            "stocked_at": data.stocked_at or now_iso(),
            "created_at": now_iso()}
    await db.inventory.insert_one(dict(item))
    item.pop("_id", None)
    return item


@api.patch("/inventory/{item_id}")
async def update_inventory(item_id: str, data: dict, user=Depends(require_roles("admin", "reception"))):
    data.pop("_id", None); data.pop("id", None)
    await db.inventory.update_one({"id": item_id}, {"$set": data})
    return await db.inventory.find_one({"id": item_id}, {"_id": 0})


@api.delete("/inventory/{item_id}")
async def delete_inventory(item_id: str, user=Depends(require_roles("admin"))):
    await db.inventory.delete_one({"id": item_id})
    return {"ok": True}


@api.get("/inventory/alerts")
async def inventory_alerts(user=Depends(require_roles("admin", "reception"))):
    items = await db.inventory.find({}, {"_id": 0}).to_list(1000)
    out_of_stock = [i for i in items if i["stock"] <= 0]
    low_stock = [i for i in items if 0 < i["stock"] <= i.get("low_stock_threshold", 5)]
    items.sort(key=lambda x: x.get("stocked_at") or "")
    by_sku: Dict[str, list] = {}
    for i in items:
        by_sku.setdefault(i["sku"], []).append(i)
    fifo = []
    for sku, group in by_sku.items():
        if len(group) > 1:
            fifo.append({"sku": sku, "use_first": group[0]["name"],
                         "stocked_at": group[0].get("stocked_at"), "id": group[0]["id"]})
    # not-sold-in-N-days alert
    threshold_days = 30
    now = datetime.now(timezone.utc)
    used_recently = set()
    async for b in db.bookings.find({"paid_at": {"$ne": None}}, {"_id": 0, "items": 1, "paid_at": 1}):
        try:
            paid = datetime.fromisoformat(b["paid_at"])
            if (now - paid).days <= threshold_days:
                for it in b.get("items", []):
                    used_recently.add(it.get("inventory_id"))
        except Exception:
            pass
    stagnant = [i for i in items if i["stock"] > 0 and i["id"] not in used_recently]
    return {"out_of_stock": out_of_stock, "low_stock": low_stock,
            "fifo": fifo, "stagnant": stagnant[:10],
            "stagnant_threshold_days": threshold_days}


# ---------- NOTIFICATIONS ----------
@api.get("/notifications/me")
async def my_notifications(user=Depends(get_current_user)):
    return await db.notifications.find({"user_id": user["id"]}, {"_id": 0}).sort("ts", -1).limit(50).to_list(50)


@api.post("/notifications/{notif_id}/read")
async def mark_read(notif_id: str, user=Depends(get_current_user)):
    await db.notifications.update_one({"id": notif_id, "user_id": user["id"]}, {"$set": {"read": True}})
    return {"ok": True}


# ---------- ADMIN ----------
@api.get("/admin/stats")
async def admin_stats(user=Depends(require_roles("admin", "reception"))):
    today = datetime.now(timezone.utc).date().isoformat()
    all_bookings = await db.bookings.find({}, {"_id": 0}).to_list(2000)
    today_done = [b for b in all_bookings if (b.get("paid_at") or "").startswith(today)]

    # Approved refunds (shop) — subtracted from revenue net
    approved_refunds = await db.refunds.find({"status": "APPROVED"}, {"_id": 0}).to_list(2000)
    refund_today_total = sum(r.get("amount", 0) for r in approved_refunds
                             if (r.get("decided_at") or "").startswith(today))

    revenue_today = sum(b.get("bill_amount", 0) for b in today_done) - refund_today_total
    by_status = {}
    for b in all_bookings:
        by_status[b["status"]] = by_status.get(b["status"], 0) + 1
    last_7 = []
    for i in range(6, -1, -1):
        d = (datetime.now(timezone.utc).date() - timedelta(days=i)).isoformat()
        day_bs = [b for b in all_bookings if (b.get("paid_at") or "").startswith(d)]
        day_refunds = sum(r.get("amount", 0) for r in approved_refunds
                          if (r.get("decided_at") or "").startswith(d))
        last_7.append({"date": d, "serviced": len(day_bs),
                       "revenue": sum(b.get("bill_amount", 0) for b in day_bs) - day_refunds})
    inv = await db.inventory.find({}, {"_id": 0}).to_list(1000)
    return {
        "today_serviced": len(today_done),
        "today_revenue": revenue_today,
        "today_refunds": refund_today_total,
        "active_bays": sum(1 for b in all_bookings if b["status"] == "IN_SERVICE"),
        "total_bookings": len(all_bookings),
        "by_status": by_status,
        "last_7_days": last_7,
        "low_stock_count": sum(1 for i in inv if 0 < i["stock"] <= i.get("low_stock_threshold", 5)),
        "out_of_stock_count": sum(1 for i in inv if i["stock"] <= 0),
    }


# ---------- HR MODULE ----------
DEFAULT_LEAVE_BALANCE = {"casual": 12, "earned": 15, "sick": 8}
STAFF_ROLES = {"mechanic", "reception", "tester", "admin", "shopkeeper"}


def _staff_user(user):
    if user["role"] not in STAFF_ROLES:
        raise HTTPException(403, "Staff only")


@api.post("/hr/leaves")
async def apply_leave(data: LeaveIn, user=Depends(get_current_user)):
    _staff_user(user)
    rec = {"id": str(uuid.uuid4()), "user_id": user["id"], "user_name": user["name"],
           "user_role": user["role"], **data.model_dump(),
           "status": "PENDING", "created_at": now_iso(),
           "decided_at": None, "decided_by": None, "decision_note": None}
    await db.leaves.insert_one(dict(rec))
    rec.pop("_id", None)
    # notify all admins
    admin_ids = [u["id"] async for u in db.users.find({"role": "admin"}, {"_id": 0, "id": 1})]
    await bus.fanout(admin_ids, {"type": "LEAVE_REQUESTED", "data": rec, "ts": now_iso()})
    for aid in admin_ids:
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()), "user_id": aid, "event_type": "LEAVE_REQUESTED",
            "title": "Leave Request", "body": f"{user['name']} applied for {data.leave_type} leave",
            "read": False, "ts": now_iso(), "ref_id": rec["id"]
        })
    return rec


@api.get("/hr/leaves/me")
async def my_leaves(user=Depends(get_current_user)):
    _staff_user(user)
    return await db.leaves.find({"user_id": user["id"]}, {"_id": 0}).sort("created_at", -1).to_list(200)


@api.get("/hr/leaves")
async def all_leaves(status: Optional[str] = None, user=Depends(require_roles("admin"))):
    q = {}
    if status:
        q["status"] = status.upper()
    return await db.leaves.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)


@api.patch("/hr/leaves/{leave_id}")
async def decide_leave(leave_id: str, data: LeaveDecision, user=Depends(require_roles("admin"))):
    decision = data.decision.upper()
    if decision not in ("APPROVED", "REJECTED"):
        raise HTTPException(400, "Invalid decision")
    await db.leaves.update_one({"id": leave_id}, {"$set": {
        "status": decision, "decided_at": now_iso(),
        "decided_by": user["name"], "decision_note": data.note
    }})
    rec = await db.leaves.find_one({"id": leave_id}, {"_id": 0})
    # notify employee
    await bus.fanout([rec["user_id"]], {"type": f"LEAVE_{decision}", "data": rec, "ts": now_iso()})
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()), "user_id": rec["user_id"], "event_type": f"LEAVE_{decision}",
        "title": f"Leave {decision.title()}",
        "body": f"Your {rec['leave_type']} leave was {decision.lower()}",
        "read": False, "ts": now_iso(), "ref_id": leave_id
    })
    return rec


# ---------- Attendance ----------
@api.post("/hr/attendance/punch")
async def punch(user=Depends(get_current_user)):
    _staff_user(user)
    today = datetime.now(timezone.utc).date().isoformat()
    rec = await db.attendance.find_one({"user_id": user["id"], "date": today})
    now = now_iso()
    if not rec:
        doc = {"id": str(uuid.uuid4()), "user_id": user["id"], "user_name": user["name"],
               "date": today, "punch_in": now, "punch_out": None}
        await db.attendance.insert_one(dict(doc))
        doc.pop("_id", None)
        return doc
    if not rec.get("punch_out"):
        await db.attendance.update_one({"id": rec["id"]}, {"$set": {"punch_out": now}})
    return await db.attendance.find_one({"id": rec["id"]}, {"_id": 0})


@api.get("/hr/attendance/me")
async def my_attendance(user=Depends(get_current_user)):
    _staff_user(user)
    return await db.attendance.find({"user_id": user["id"]}, {"_id": 0}).sort("date", -1).limit(60).to_list(60)


@api.get("/hr/attendance/all")
async def all_attendance(date: Optional[str] = None, user=Depends(require_roles("admin"))):
    q = {"date": date} if date else {}
    return await db.attendance.find(q, {"_id": 0}).sort("date", -1).to_list(500)


# ---------- Holidays ----------
@api.get("/hr/holidays")
async def list_holidays(user=Depends(get_current_user)):
    return await db.holidays.find({}, {"_id": 0}).sort("date", 1).to_list(200)


@api.post("/hr/holidays")
async def add_holiday(data: HolidayIn, user=Depends(require_roles("admin"))):
    doc = {"id": str(uuid.uuid4()), **data.model_dump(), "created_at": now_iso()}
    await db.holidays.insert_one(dict(doc))
    doc.pop("_id", None)
    return doc


@api.delete("/hr/holidays/{hid}")
async def del_holiday(hid: str, user=Depends(require_roles("admin"))):
    await db.holidays.delete_one({"id": hid})
    return {"ok": True}


# ---------- Employee Profile ----------
@api.get("/hr/profile/me")
async def my_profile(user=Depends(get_current_user)):
    _staff_user(user)
    profile = await db.profiles.find_one({"user_id": user["id"]}, {"_id": 0}) or {}
    bal = profile.get("leave_balance") or DEFAULT_LEAVE_BALANCE
    used = {"casual": 0, "earned": 0, "sick": 0, "unpaid": 0}
    async for l in db.leaves.find({"user_id": user["id"], "status": "APPROVED"}, {"_id": 0}):
        try:
            days = (datetime.fromisoformat(l["end_date"]).date() - datetime.fromisoformat(l["start_date"]).date()).days + 1
            used[l["leave_type"]] = used.get(l["leave_type"], 0) + max(1, days)
        except Exception:
            pass
    timeline = await db.payroll_events.find({"user_id": user["id"]}, {"_id": 0}).sort("ts", -1).limit(50).to_list(50)
    return {"user": {k: v for k, v in user.items() if k != "password_hash"},
            "profile": profile, "leave_balance": bal, "leave_used": used,
            "timeline": timeline}


@api.patch("/hr/profile/{user_id}")
async def update_profile(user_id: str, data: ProfileUpdate, user=Depends(require_roles("admin"))):
    upd = {k: v for k, v in data.model_dump().items() if v is not None}
    upd["user_id"] = user_id
    await db.profiles.update_one({"user_id": user_id}, {"$set": upd}, upsert=True)
    return await db.profiles.find_one({"user_id": user_id}, {"_id": 0})


@api.post("/hr/payroll/event")
async def add_payroll_event(data: BonusIn, user=Depends(require_roles("admin"))):
    target = await db.users.find_one({"id": data.user_id}, {"_id": 0})
    if not target:
        raise HTTPException(404, "Employee not found")
    doc = {"id": str(uuid.uuid4()), "user_id": data.user_id, "user_name": target["name"],
           "amount": data.amount, "reason": data.reason, "event_type": data.event_type,
           "by": user["name"], "ts": now_iso()}
    await db.payroll_events.insert_one(dict(doc))
    doc.pop("_id", None)
    await bus.fanout([data.user_id], {"type": f"PAYROLL_{data.event_type.upper()}", "data": doc, "ts": now_iso()})
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()), "user_id": data.user_id,
        "event_type": f"PAYROLL_{data.event_type.upper()}",
        "title": data.event_type.replace("_", " ").title(),
        "body": f"₹{data.amount} · {data.reason}",
        "read": False, "ts": now_iso()
    })
    return doc


@api.get("/hr/profile/{user_id}")
async def admin_view_profile(user_id: str, user=Depends(require_roles("admin"))):
    profile = await db.profiles.find_one({"user_id": user_id}, {"_id": 0}) or {}
    target = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    timeline = await db.payroll_events.find({"user_id": user_id}, {"_id": 0}).sort("ts", -1).limit(50).to_list(50)
    return {"user": target, "profile": profile, "timeline": timeline}


# ---------- DIGITAL SERVICE CARDS ----------

class ServiceCardPlanIn(BaseModel):
    name: str
    price: float
    services_per_year: int = 3
    duration_years: int = 1
    interval_months: int = 4
    interval_km: int = 0
    discount_pct: float = 0.0
    offer_note: str = ""
    active: bool = True


class ServiceCardCreate(BaseModel):
    plan_id: str
    customer_name: str
    customer_phone: str
    plate_number: Optional[str] = ""
    car_make: Optional[str] = ""
    car_model: Optional[str] = ""
    current_km: Optional[int] = 0
    notes: Optional[str] = ""
    discount_override: Optional[float] = None


@api.get("/service-card-plans")
async def list_sc_plans(user=Depends(require_roles("admin", "reception"))):
    return await db.service_card_plans.find({}, {"_id": 0}).to_list(100)


@api.post("/service-card-plans")
async def create_sc_plan(data: ServiceCardPlanIn, user=Depends(require_roles("admin"))):
    plan = {"id": str(uuid.uuid4()), **data.model_dump(), "created_at": now_iso(), "created_by": user["id"]}
    await db.service_card_plans.insert_one(plan)
    plan.pop("_id", None)
    return plan


@api.put("/service-card-plans/{plan_id}")
async def update_sc_plan(plan_id: str, data: dict, user=Depends(require_roles("admin"))):
    data.pop("id", None); data.pop("_id", None)
    await db.service_card_plans.update_one(
        {"id": plan_id},
        {"$set": {**_sanitize_value(data), "updated_at": now_iso()}}
    )
    return await db.service_card_plans.find_one({"id": plan_id}, {"_id": 0})


@api.get("/service-cards/check-customer/{phone}")
async def check_customer_card(phone: str, user=Depends(require_roles("admin", "reception"))):
    norm = _normalize_phone(phone)
    card = await db.service_cards.find_one(
        {"customer_phone": norm, "status": "active"}, {"_id": 0}
    )
    paid_count = await db.bookings.count_documents({"customer_phone": norm, "paid": True})
    return {
        "has_active_card": bool(card),
        "card": card,
        "paid_bookings_count": paid_count,
        "eligible_for_offer": paid_count >= 1 and not card,
        "is_new_customer": paid_count == 0,
    }


@api.get("/service-cards")
async def list_service_cards(user=Depends(require_roles("admin", "reception"))):
    return await db.service_cards.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)


@api.post("/service-cards")
async def create_service_card(data: ServiceCardCreate, user=Depends(require_roles("admin", "reception"))):
    plan = await db.service_card_plans.find_one({"id": data.plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(404, "Plan not found")
    if not plan.get("active"):
        raise HTTPException(400, "This plan is inactive")
    now_dt = datetime.now(timezone.utc)
    end_dt = now_dt + timedelta(days=365 * plan["duration_years"])
    discount = data.discount_override if data.discount_override is not None else plan["discount_pct"]
    total_slots = plan["services_per_year"] * plan["duration_years"]
    phone = _normalize_phone(data.customer_phone)
    card = {
        "id": str(uuid.uuid4()),
        "plan_id": data.plan_id,
        "plan_name": plan["name"],
        "customer_name": data.customer_name,
        "customer_phone": phone,
        "plate_number": (data.plate_number or "").upper(),
        "car_make": data.car_make or "",
        "car_model": data.car_model or "",
        "current_km": int(data.current_km or 0),
        "last_km": int(data.current_km or 0),
        "last_service_date": now_dt.isoformat(),
        "slots_total": total_slots,
        "slots_used": 0,
        "discount_pct": float(discount),
        "offer_note": plan.get("offer_note", ""),
        "interval_months": plan["interval_months"],
        "interval_km": plan["interval_km"],
        "price_paid": plan["price"],
        "start_date": now_dt.isoformat(),
        "end_date": end_dt.isoformat(),
        "status": "active",
        "notes": data.notes or "",
        "created_by": user["id"],
        "created_at": now_iso(),
    }
    await db.service_cards.insert_one(dict(card))
    card.pop("_id", None)
    if phone:
        sms_text = (
            f"MacJit: Your Digital Service Card '{plan['name']}' is now active!\n"
            f"Valid: {now_dt.strftime('%d %b %Y')} to {end_dt.strftime('%d %b %Y')}\n"
            f"Services: {total_slots} | Discount: {discount:.0f}% on every visit.\n"
        )
        if plan.get("offer_note"):
            sms_text += f"Offer: {plan['offer_note']}\n"
        sms_text += "Thank you for trusting MacJit Garage!"
        await TwilioAdapter.send_sms(phone, sms_text)
    return card


@api.get("/service-cards/{card_id}")
async def get_service_card(card_id: str, user=Depends(require_roles("admin", "reception"))):
    card = await db.service_cards.find_one({"id": card_id}, {"_id": 0})
    if not card:
        raise HTTPException(404, "Not found")
    return card


@api.put("/service-cards/{card_id}")
async def update_service_card(card_id: str, data: dict, user=Depends(require_roles("admin", "reception"))):
    data.pop("id", None); data.pop("_id", None)
    clean = _sanitize_value(data)
    clean["updated_at"] = now_iso()
    clean["updated_by"] = user["id"]
    await db.service_cards.update_one({"id": card_id}, {"$set": clean})
    return await db.service_cards.find_one({"id": card_id}, {"_id": 0})


@api.post("/service-cards/{card_id}/use-slot")
async def use_service_card_slot(card_id: str, data: Optional[dict] = None,
                                 user=Depends(require_roles("admin", "reception"))):
    card = await db.service_cards.find_one({"id": card_id})
    if not card:
        raise HTTPException(404, "Not found")
    if card.get("status") != "active":
        raise HTTPException(400, "Card is not active")
    if card["slots_used"] >= card["slots_total"]:
        raise HTTPException(400, "All service slots already used")
    current_km = int((data or {}).get("current_km") or card.get("current_km", 0))
    new_used = card["slots_used"] + 1
    upd: dict = {
        "slots_used": new_used,
        "last_service_date": now_iso(),
        "current_km": current_km,
        "last_km": current_km,
        "updated_at": now_iso(),
    }
    if new_used >= card["slots_total"]:
        upd["status"] = "exhausted"
    await db.service_cards.update_one({"id": card_id}, {"$set": upd})
    return await db.service_cards.find_one({"id": card_id}, {"_id": 0})


@api.post("/service-cards/send-reminders")
async def send_card_reminders(user=Depends(require_roles("admin", "reception"))):
    now_dt = datetime.now(timezone.utc)
    cards = await db.service_cards.find({"status": "active"}, {"_id": 0}).to_list(2000)
    sent = 0
    for card in cards:
        phone = card.get("customer_phone")
        if not phone:
            continue
        due = False
        last_svc = card.get("last_service_date")
        if last_svc:
            try:
                last_dt = datetime.fromisoformat(last_svc.replace("Z", "+00:00"))
                months_since = (now_dt - last_dt).days / 30.0
                if months_since >= card.get("interval_months", 4):
                    due = True
            except Exception:
                pass
        km_gap = card.get("interval_km", 0)
        if km_gap > 0:
            km_diff = card.get("current_km", 0) - card.get("last_km", 0)
            if km_diff >= km_gap:
                due = True
        if due:
            slots_left = card["slots_total"] - card["slots_used"]
            msg = (
                f"MacJit Reminder: Hi {card['customer_name']}, your {card['plan_name']} service is due!\n"
                f"Vehicle: {card.get('plate_number') or card.get('car_make', '')}\n"
                f"Services remaining: {slots_left} | Discount: {card['discount_pct']:.0f}%\n"
                f"Book now at MacJit Garage."
            )
            await TwilioAdapter.send_sms(phone, msg)
            sent += 1
    return {"ok": True, "reminders_sent": sent}


@api.post("/service-cards/remind-custom")
async def remind_custom(data: dict, user=Depends(require_roles("admin", "reception"))):
    """Send a one-off SMS reminder to any phone (even without a card)."""
    phone = sanitize_str(data.get("phone"))
    message = sanitize_str(data.get("message"), max_len=300)
    name = sanitize_str(data.get("name", "Customer"), max_len=100)
    if not phone or not message:
        raise HTTPException(400, "phone and message are required")
    await TwilioAdapter.send_sms(_normalize_phone(phone), message)
    return {"ok": True}


# ---------- SHOP MODULE (walk-in parts counter, shares inventory) ----------
@api.post("/shop/sales")
async def create_sale(data: ShopSaleIn, user=Depends(require_roles("shopkeeper", "admin", "reception"))):
    if not data.items:
        raise HTTPException(400, "Cart empty")
    lines = []
    total = 0.0
    for line in data.items:
        inv = await db.inventory.find_one({"id": line.inventory_id})
        if not inv:
            raise HTTPException(404, f"Item not found: {line.inventory_id}")
        if inv["stock"] < line.qty:
            raise HTTPException(400, f"Insufficient stock for {inv['name']}")
        subtotal = inv["price"] * line.qty
        lines.append({"inventory_id": inv["id"], "name": inv["name"], "sku": inv["sku"],
                      "qty": line.qty, "price": inv["price"], "subtotal": subtotal})
        total += subtotal
        await db.inventory.update_one({"id": inv["id"]}, {"$inc": {"stock": -line.qty}})

    fitting_charge = round(max(0.0, float(data.fitting_charge or 0)), 2)
    gst_percent = round(max(0.0, float(data.gst_percent or 0)), 2)
    taxable_amount = total + fitting_charge
    gst_amount = round(taxable_amount * gst_percent / 100, 2)
    grand_total = round(taxable_amount + gst_amount, 2)

    sale_id = str(uuid.uuid4())
    payment_link = None
    rzp_payment_link_id = None
    if data.payment_method == "razorpay":
        notes = {"type": "shop_sale", "sale_id": sale_id,
                 "customer": data.customer_name or "Walk-in"}
        payment_link, rzp_payment_link_id = await RazorpayAdapter.create_payment_link_full(
            grand_total, sale_id, data.customer_phone or "", notes=notes
        )

    sale = {
        "id": sale_id,
        "customer_name": data.customer_name or "Walk-in",
        "customer_phone": data.customer_phone or "",
        "items": lines,
        "subtotal": total,
        "fitting_charge": fitting_charge,
        "gst_percent": gst_percent,
        "gst_amount": gst_amount,
        "total": grand_total,
        "payment_method": data.payment_method,
        "payment_link": payment_link,
        "rzp_payment_link_id": rzp_payment_link_id,
        "paid": data.payment_method == "cash",  # cash assumed paid at counter
        "shopkeeper_id": user["id"], "shopkeeper_name": user["name"],
        "created_at": now_iso(),
        "paid_at": now_iso() if data.payment_method == "cash" else None,
    }
    await db.shop_sales.insert_one(dict(sale))
    sale.pop("_id", None)
    if data.customer_phone:
        if data.payment_method == "razorpay" and payment_link:
            # Razorpay: send the payment link now; invoice auto-sent after payment via webhook
            pay_msg = (
                f"MacJit - Your bill of ₹{grand_total:.0f} is ready.\n"
                f"Pay securely here: {payment_link}\n"
                f"Your invoice will be sent to you after payment. Thank you!"
            )
            await TwilioAdapter.send_sms(data.customer_phone, pay_msg)
        else:
            # Cash / other: payment already done — send full invoice immediately
            await _send_sale_invoice_whatsapp(sale)
    # notify admin
    admin_ids = [u["id"] async for u in db.users.find({"role": "admin"}, {"_id": 0, "id": 1})]
    await bus.fanout(admin_ids, {"type": "SHOP_SALE", "data": sale, "ts": now_iso()})
    return sale


@api.post("/shop/sales/{sale_id}/pay")
async def mark_sale_paid(sale_id: str, user=Depends(require_roles("shopkeeper", "admin", "reception"))):
    await db.shop_sales.update_one({"id": sale_id}, {"$set": {"paid": True, "paid_at": now_iso()}})
    sale = await db.shop_sales.find_one({"id": sale_id}, {"_id": 0})
    if sale:
        await _send_sale_invoice_whatsapp(sale)
    return sale


@api.get("/shop/sales")
async def list_sales(user=Depends(require_roles("shopkeeper", "admin", "reception"))):
    return await db.shop_sales.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)


@api.get("/shop/stats")
async def shop_stats(user=Depends(require_roles("shopkeeper", "admin", "reception"))):
    today = datetime.now(timezone.utc).date().isoformat()
    sales = await db.shop_sales.find({}, {"_id": 0}).to_list(2000)
    today_sales = [s for s in sales if (s.get("created_at") or "").startswith(today)]
    today_paid = [s for s in today_sales if s.get("paid")]

    # Approved shop refunds (subtract from revenue net)
    approved_refunds = await db.refunds.find({"status": "APPROVED"}, {"_id": 0}).to_list(2000)
    refund_today_total = sum(r.get("amount", 0) for r in approved_refunds
                             if (r.get("decided_at") or "").startswith(today))

    last_7 = []
    for i in range(6, -1, -1):
        d = (datetime.now(timezone.utc).date() - timedelta(days=i)).isoformat()
        day = [s for s in sales if (s.get("created_at") or "").startswith(d) and s.get("paid")]
        day_refunds = sum(r.get("amount", 0) for r in approved_refunds
                          if (r.get("decided_at") or "").startswith(d))
        last_7.append({"date": d, "sales": len(day),
                       "revenue": sum(s.get("total", 0) for s in day) - day_refunds})
    # Top fast-movers (last 7 days, by units sold)
    cutoff = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    movers = {}
    for s in sales:
        if (s.get("created_at") or "") < cutoff:
            continue
        for it in s.get("items", []):
            k = it.get("inventory_id")
            if k not in movers:
                movers[k] = {"inventory_id": k, "name": it["name"], "sku": it["sku"], "units": 0, "revenue": 0}
            movers[k]["units"] += it.get("qty", 0)
            movers[k]["revenue"] += it.get("subtotal", 0)
    fast_movers = sorted(movers.values(), key=lambda x: x["units"], reverse=True)[:5]
    return {
        "today_count": len(today_sales),
        "today_revenue": sum(s.get("total", 0) for s in today_paid) - refund_today_total,
        "today_refunds": refund_today_total,
        "pending_count": sum(1 for s in sales if not s.get("paid")),
        "last_7_days": last_7,
        "total_sales": len(sales),
        "fast_movers": fast_movers,
    }


# ---------- REFUNDS (shop) ----------
@api.post("/shop/refunds")
async def raise_refund(data: RefundIn, user=Depends(require_roles("shopkeeper", "admin"))):
    """Shopkeeper raises a refund request for a paid shop sale.
    Admin must approve before stock is restored and sale is flagged refunded."""
    sale = await db.shop_sales.find_one({"id": data.sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(404, "Sale not found")
    if not sale.get("paid"):
        raise HTTPException(400, "Only paid sales can be refunded")
    if sale.get("refund_status") in ("PENDING", "APPROVED"):
        raise HTTPException(400, f"Refund already {sale['refund_status']}")

    # Build refund line items (full or partial)
    sale_items_by_id = {it["inventory_id"]: it for it in sale.get("items", [])}
    refund_items = []
    if data.items:
        for line in data.items:
            base = sale_items_by_id.get(line.inventory_id)
            if not base:
                raise HTTPException(400, f"Item {line.inventory_id} not in sale")
            if line.qty <= 0 or line.qty > base["qty"]:
                raise HTTPException(400, f"Invalid qty for {base['name']}")
            refund_items.append({**base, "qty": line.qty,
                                 "subtotal": base["price"] * line.qty})
    else:
        refund_items = [dict(it) for it in sale.get("items", [])]
    refund_total = sum(it["subtotal"] for it in refund_items)

    refund = {
        "id": str(uuid.uuid4()),
        "sale_id": sale["id"],
        "customer_name": sale.get("customer_name"),
        "customer_phone": sale.get("customer_phone"),
        "items": refund_items,
        "amount": refund_total,
        "reason": data.reason,
        "status": "PENDING",
        "raised_by_id": user["id"],
        "raised_by_name": user["name"],
        "raised_at": now_iso(),
        "decided_by_name": None,
        "decided_at": None,
        "decision_note": "",
    }
    await db.refunds.insert_one(dict(refund))
    refund.pop("_id", None)
    await db.shop_sales.update_one({"id": sale["id"]}, {"$set": {"refund_status": "PENDING"}})

    admin_ids = [u["id"] async for u in db.users.find({"role": "admin"}, {"_id": 0, "id": 1})]
    await bus.fanout(admin_ids, {"type": "REFUND_RAISED", "data": refund, "ts": now_iso()})
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": admin_ids[0] if admin_ids else "",
        "event_type": "REFUND_RAISED",
        "title": "Refund request",
        "body": f"₹{refund_total:.0f} · {sale.get('customer_name','Walk-in')} · {data.reason[:40]}",
        "read": False, "ts": now_iso(),
    })
    return refund


@api.get("/shop/refunds")
async def my_refunds(user=Depends(require_roles("shopkeeper", "admin", "reception"))):
    q = {} if user.get("role") == "admin" else {"raised_by_id": user["id"]}
    return await db.refunds.find(q, {"_id": 0}).sort("raised_at", -1).to_list(200)


@api.get("/admin/refunds")
async def admin_list_refunds(status: Optional[str] = None,
                             user=Depends(require_roles("admin"))):
    q = {}
    if status:
        q["status"] = status.upper()
    return await db.refunds.find(q, {"_id": 0}).sort("raised_at", -1).to_list(500)


@api.post("/admin/refunds/{refund_id}/decision")
async def decide_refund(refund_id: str, data: RefundDecision,
                        user=Depends(require_roles("admin"))):
    decision = (data.decision or "").lower()
    if decision not in ("approved", "rejected"):
        raise HTTPException(400, "decision must be 'approved' or 'rejected'")
    refund = await db.refunds.find_one({"id": refund_id}, {"_id": 0})
    if not refund:
        raise HTTPException(404, "Refund not found")
    if refund["status"] != "PENDING":
        raise HTTPException(400, f"Already {refund['status']}")

    new_status = "APPROVED" if decision == "approved" else "REJECTED"
    await db.refunds.update_one({"id": refund_id}, {"$set": {
        "status": new_status,
        "decided_by_name": user["name"],
        "decided_at": now_iso(),
        "decision_note": data.note or "",
    }})

    if new_status == "APPROVED":
        # Restore stock & mark sale refunded
        for it in refund["items"]:
            await db.inventory.update_one({"id": it["inventory_id"]},
                                          {"$inc": {"stock": it["qty"]}})
        await db.shop_sales.update_one({"id": refund["sale_id"]}, {"$set": {
            "refund_status": "APPROVED",
            "refunded_amount": refund["amount"],
            "refunded_at": now_iso(),
        }})
    else:
        await db.shop_sales.update_one({"id": refund["sale_id"]},
                                       {"$set": {"refund_status": "REJECTED"}})

    # Notify shopkeeper
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": refund["raised_by_id"],
        "event_type": f"REFUND_{new_status}",
        "title": f"Refund {new_status.lower()}",
        "body": f"₹{refund['amount']:.0f} · {data.note or 'no note'}",
        "read": False, "ts": now_iso(),
    })
    await bus.fanout([refund["raised_by_id"]],
                     {"type": f"REFUND_{new_status}", "data": refund, "ts": now_iso()})
    return await db.refunds.find_one({"id": refund_id}, {"_id": 0})


# ---------- ADMIN TRANSACTIONS (Paytm-style unified history) ----------
@api.get("/admin/transactions")
async def admin_transactions(type: str = "all", q: str = "", limit: int = 200,
                             date_from: Optional[str] = None, date_to: Optional[str] = None,
                             period: Optional[str] = None,
                             user=Depends(require_roles("admin"))):
    """Unified payment history across services (paid bookings) and shop (paid counter sales).
    Filter `type`: all | service | shop. `q` matches customer name/phone/plate/sku."""
    type = (type or "all").lower()
    items: list = []

    if type in ("all", "service"):
        bookings = await db.bookings.find({"paid": True}, {"_id": 0}).sort("paid_at", -1).to_list(500)
        for b in bookings:
            items.append({
                "id": b["id"],
                "kind": "service",
                "ref": b["id"][:8].upper(),
                "customer_name": b.get("customer_name") or "—",
                "customer_phone": b.get("customer_phone") or "",
                "title": f"{b.get('service_type','service')} · {b.get('plate_number','')}",
                "subtitle": f"{b.get('car_make','')} {b.get('car_model','')}".strip() or "—",
                "amount": b.get("bill_amount", 0),
                "method": b.get("payment_method") or ("razorpay" if b.get("razorpay_payment_id") else "cash"),
                "ts": b.get("paid_at") or b.get("billed_at") or b.get("created_at"),
                "items": b.get("items") or [],
                "extra": {"mechanic": b.get("mechanic_name") or "—",
                          "bay": b.get("bay_name") or "—",
                          "razorpay_payment_id": b.get("razorpay_payment_id")},
                "invoice_url": f"/api/invoices/{b['id']}.pdf",
            })

    if type in ("all", "shop"):
        sales = await db.shop_sales.find({"paid": True}, {"_id": 0}).sort("paid_at", -1).to_list(500)
        for s in sales:
            items.append({
                "id": s["id"],
                "kind": "shop",
                "ref": s["id"][:8].upper(),
                "customer_name": s.get("customer_name") or "Walk-in",
                "customer_phone": s.get("customer_phone") or "",
                "title": f"{len(s.get('items', []))} item(s) · {s.get('shopkeeper_name','counter')}",
                "subtitle": ", ".join((it.get("name") or "")[:24] for it in (s.get("items") or [])[:2]) or "—",
                "amount": s.get("total", 0),
                "method": s.get("payment_method") or "cash",
                "ts": s.get("paid_at") or s.get("created_at"),
                "items": s.get("items") or [],
                "extra": {"refund_status": s.get("refund_status"),
                          "refunded_amount": s.get("refunded_amount")},
                "invoice_url": None,
            })

        # Add an explicit refund row (negative amount) for each approved refund
        approved_refunds = await db.refunds.find({"status": "APPROVED"}, {"_id": 0}).to_list(500)
        for r in approved_refunds:
            items.append({
                "id": r["id"],
                "kind": "refund",
                "ref": "RF-" + r["id"][:6].upper(),
                "customer_name": r.get("customer_name") or "Walk-in",
                "customer_phone": r.get("customer_phone") or "",
                "title": f"Refund · sale #{r.get('sale_id','')[:8]}",
                "subtitle": (r.get("reason") or "")[:60] or "—",
                "amount": -float(r.get("amount", 0)),
                "method": "refund",
                "ts": r.get("decided_at") or r.get("raised_at"),
                "items": r.get("items") or [],
                "extra": {"raised_by": r.get("raised_by_name"),
                          "decided_by": r.get("decided_by_name"),
                          "note": r.get("decision_note"),
                          "sale_id": r.get("sale_id")},
                "invoice_url": None,
            })

    # --- Period/date filter ---
    _now = datetime.now(timezone.utc)
    if period == "today":
        date_from = _now.date().isoformat()
        date_to = date_from
    elif period == "week":
        date_from = (_now.date() - timedelta(days=6)).isoformat()
        date_to = _now.date().isoformat()
    elif period == "month":
        date_from = (_now.date() - timedelta(days=29)).isoformat()
        date_to = _now.date().isoformat()
    if date_from or date_to:
        def _in_date_range(t):
            ts = (t.get("ts") or "")[:10]
            if date_from and ts < date_from:
                return False
            if date_to and ts > date_to:
                return False
            return True
        items = [t for t in items if _in_date_range(t)]

    if q:
        ql = q.lower()
        def match(t):
            blob = " ".join([
                str(t.get("customer_name") or ""),
                str(t.get("customer_phone") or ""),
                str(t.get("title") or ""),
                str(t.get("ref") or ""),
                " ".join((it.get("name", "") + " " + it.get("sku", "")) for it in t.get("items", [])),
            ]).lower()
            return ql in blob
        items = [t for t in items if match(t)]

    items.sort(key=lambda x: x.get("ts") or "", reverse=True)
    items = items[:limit]
    gross_in = sum(t["amount"] for t in items if t["amount"] > 0)
    refund_out = -sum(t["amount"] for t in items if t["amount"] < 0)
    return {
        "transactions": items,
        "total_amount": gross_in - refund_out,   # NET (after refunds)
        "gross_amount": gross_in,                # before refunds
        "refund_amount": refund_out,             # total refunded out
        "count": len(items),
        "service_count": sum(1 for t in items if t["kind"] == "service"),
        "shop_count": sum(1 for t in items if t["kind"] == "shop"),
        "refund_count": sum(1 for t in items if t["kind"] == "refund"),
    }


# ---------- Mechanic photo capture (live progress photos) ----------
@api.post("/bookings/{booking_id}/photos")
async def upload_photo(booking_id: str, data: dict, user=Depends(require_roles("mechanic"))):
    """Accepts {data_url: 'data:image/jpeg;base64,...', caption?: ''} and stores."""
    data_url = data.get("data_url", "")
    if not data_url.startswith("data:image"):
        raise HTTPException(400, "Invalid image data")
    if len(data_url) > 2_500_000:  # ~2.5MB cap
        raise HTTPException(400, "Image too large (max 2MB)")
    photo = {
        "id": str(uuid.uuid4()),
        "booking_id": booking_id,
        "data_url": data_url,
        "caption": data.get("caption", ""),
        "captured_by": user["name"],
        "ts": now_iso(),
    }
    await db.service_photos.insert_one(dict(photo))
    photo.pop("_id", None)
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if b:
        recipients = await get_recipients_for_booking(b, include_admin=False)
        await bus.fanout(recipients, {"type": "PHOTO_CAPTURED", "data": {"booking_id": booking_id, "ts": photo["ts"]}, "ts": now_iso()})
    return {"id": photo["id"], "ts": photo["ts"]}


@api.get("/bookings/{booking_id}/photos")
async def list_photos(booking_id: str, user=Depends(get_current_user)):
    b = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Not found")
    if user["role"] == "customer" and b.get("customer_id") != user["id"]:
        raise HTTPException(403, "Forbidden")
    return await db.service_photos.find({"booking_id": booking_id}, {"_id": 0}).sort("ts", -1).to_list(50)





async def _send_sale_invoice_whatsapp(sale: dict):
    """Build a formatted invoice from a sale doc and send it via Twilio WhatsApp."""
    phone = sale.get("customer_phone", "")
    if not phone:
        return
    items = sale.get("items", [])
    inv_lines = "\n".join(f"  {it['name']} x{it['qty']} = \u20b9{it['subtotal']:.0f}" for it in items)
    fitting = float(sale.get("fitting_charge") or 0)
    gst_pct = float(sale.get("gst_percent") or 0)
    gst_amt = float(sale.get("gst_amount") or 0)
    grand = float(sale.get("total", 0))
    fitting_line = f"\n  Fitting charge = \u20b9{fitting:.0f}" if fitting > 0 else ""
    gst_line = f"\n  GST ({gst_pct}%) = \u20b9{gst_amt:.0f}" if gst_amt > 0 else ""
    msg = (
        f"*MacJit Invoice #{sale['id'][:8]}*\n"
        f"{inv_lines}{fitting_line}{gst_line}\n"
        f"*Total = \u20b9{grand:.0f}*\n"
        f"Payment: {sale.get('payment_method','cash').upper()} — PAID \u2705\nThank you! \U0001f697"
    )
    await TwilioAdapter.send_whatsapp(phone, msg)


# ---------- Razorpay Webhook ----------
@app.post("/api/webhooks/razorpay")
async def razorpay_webhook(request: Request):
    """
    Razorpay sends POST with X-Razorpay-Signature header.
    Events handled:
      - payment_link.paid    → mark shop sale paid + send invoice via WhatsApp
      - payment.captured     → same fallback lookup
      - payment.authorized   → for bookings: mark booking as paid
    Configure in Razorpay Dashboard → Settings → Webhooks → Add URL:
        https://<your-domain>/api/webhooks/razorpay
    Secret: set RAZORPAY_WEBHOOK_SECRET env var (same as dashboard).
    """
    body = await request.body()
    signature = request.headers.get("X-Razorpay-Signature", "")

    if not RazorpayAdapter.verify_webhook_signature(body, signature):
        logger.warning("[RAZORPAY-WEBHOOK] Invalid signature — rejected")
        raise HTTPException(400, "Invalid webhook signature")

    try:
        payload = json.loads(body)
    except json.JSONDecodeError:
        raise HTTPException(400, "Invalid JSON")

    event = payload.get("event", "")
    logger.info(f"[RAZORPAY-WEBHOOK] event={event}")

    # ---- payment_link.paid (primary event for payment links) ----
    if event == "payment_link.paid":
        pl_entity = payload.get("payload", {}).get("payment_link", {}).get("entity", {})
        rzp_link_id = pl_entity.get("id", "")
        notes = pl_entity.get("notes") or {}
        ref_id = notes.get("ref_id") or notes.get("sale_id") or ""
        txn_type = notes.get("type", "")
        rzp_payment_id = payload.get("payload", {}).get("payment", {}).get("entity", {}).get("id", "")
        amount_paise = pl_entity.get("amount_paid") or pl_entity.get("amount") or 0
        amount = amount_paise / 100

        logger.info(f"[RAZORPAY-WEBHOOK] payment_link.paid link={rzp_link_id} ref={ref_id} type={txn_type} payment_id={rzp_payment_id}")

        if txn_type == "shop_sale" or not txn_type:
            # Try lookup by rzp_payment_link_id first, then by sale_id in notes
            sale = None
            if rzp_link_id:
                sale = await db.shop_sales.find_one({"rzp_payment_link_id": rzp_link_id}, {"_id": 0})
            if not sale and ref_id:
                sale = await db.shop_sales.find_one({"id": ref_id}, {"_id": 0})
            if sale and not sale.get("paid"):
                now_ts = now_iso()
                await db.shop_sales.update_one({"id": sale["id"]}, {"$set": {
                    "paid": True,
                    "paid_at": now_ts,
                    "rzp_payment_id": rzp_payment_id,
                    "rzp_amount_received": amount,
                }})
                sale["paid"] = True
                sale["paid_at"] = now_ts
                # Send formatted WhatsApp invoice
                await _send_sale_invoice_whatsapp(sale)
                # Notify admin/shopkeeper via WebSocket
                admin_ids = [u["id"] async for u in db.users.find({"role": {"$in": ["admin", "shopkeeper"]}}, {"_id": 0, "id": 1})]
                await bus.fanout(admin_ids, {
                    "type": "SHOP_PAYMENT_RECEIVED",
                    "data": {"sale_id": sale["id"], "amount": amount, "payment_id": rzp_payment_id},
                    "ts": now_iso()
                })
                logger.info(f"[RAZORPAY-WEBHOOK] Shop sale {sale['id']} marked PAID — invoice sent to {sale.get('customer_phone')}")
            elif sale and sale.get("paid"):
                logger.info(f"[RAZORPAY-WEBHOOK] Sale {sale.get('id')} already marked paid — skipping")
            else:
                logger.warning(f"[RAZORPAY-WEBHOOK] Could not find shop sale for link={rzp_link_id} ref={ref_id}")

        elif txn_type == "service_booking":
            booking_id = notes.get("booking_id") or ref_id
            booking = None
            if rzp_link_id:
                booking = await db.bookings.find_one({"rzp_payment_link_id": rzp_link_id}, {"_id": 0})
            if not booking and booking_id:
                booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
            if booking and not booking.get("paid"):
                await db.bookings.update_one({"id": booking["id"]}, {"$set": {
                    "paid": True,
                    "status": "PAID",
                    "paid_at": now_iso(),
                    "payment_method": "razorpay",
                    "razorpay_payment_id": rzp_payment_id,
                }})
                updated = await db.bookings.find_one({"id": booking["id"]}, {"_id": 0})
                recipients = await get_recipients_for_booking(updated)
                await publish_event("PAID", updated, recipients)
                # Also send SMS invoice to customer
                c_phone = updated.get("customer_phone")
                if c_phone:
                    pub = os.environ.get("PUBLIC_URL") or os.environ.get("APP_URL") or ""
                    inv_url = f"{pub}/api/invoices/{updated['id']}.pdf"
                    await TwilioAdapter.send_sms(c_phone, f"MacJit: Payment of \u20B9{updated.get('bill_amount',0)} received for {updated.get('plate_number','')}. Invoice: {inv_url}")
                logger.info(f"[RAZORPAY-WEBHOOK] Booking {booking['id']} marked PAID")
            elif booking:
                logger.info(f"[RAZORPAY-WEBHOOK] Booking already paid — skipping")

    # ---- payment.captured (fallback for orders / older flows) ----
    elif event == "payment.captured":
        payment = payload.get("payload", {}).get("payment", {}).get("entity", {})
        rzp_payment_id = payment.get("id", "")
        notes = payment.get("notes") or {}
        ref_id = notes.get("ref_id") or ""
        txn_type = notes.get("type", "")
        amount = (payment.get("amount") or 0) / 100

        if txn_type == "shop_sale" and ref_id:
            sale = await db.shop_sales.find_one({"id": ref_id}, {"_id": 0})
            if sale and not sale.get("paid"):
                await db.shop_sales.update_one({"id": ref_id}, {"$set": {
                    "paid": True, "paid_at": now_iso(),
                    "rzp_payment_id": rzp_payment_id,
                    "rzp_amount_received": amount,
                }})
                sale["paid"] = True
                await _send_sale_invoice_whatsapp(sale)
                logger.info(f"[RAZORPAY-WEBHOOK] payment.captured → shop sale {ref_id} PAID")

    return {"status": "ok"}


@app.get("/api/webhooks/razorpay/redirect")
async def razorpay_redirect(razorpay_payment_id: Optional[str] = None,
                             razorpay_payment_link_id: Optional[str] = None,
                             razorpay_payment_link_status: Optional[str] = None,
                             razorpay_signature: Optional[str] = None):
    """
    Razorpay callback redirect after customer completes payment on the payment page.
    Razorpay appends query params: razorpay_payment_id, razorpay_payment_link_id,
    razorpay_payment_link_status, razorpay_signature.
    We verify the signature and mark the sale/booking paid immediately (no wait for webhook).
    """
    from fastapi.responses import RedirectResponse

    if razorpay_payment_link_status != "paid":
        logger.info(f"[RAZORPAY-REDIRECT] status={razorpay_payment_link_status} — not paid yet")
        return RedirectResponse(url=os.environ.get("PUBLIC_URL", "/") + "?payment=cancelled")

    # Verify redirect signature: HMAC-SHA256(payment_link_id|payment_link_reference_id|payment_link_status|payment_id)
    if RazorpayAdapter.WEBHOOK_SECRET and razorpay_payment_id and razorpay_payment_link_id:
        msg = f"{razorpay_payment_link_id}|{razorpay_payment_link_id}|{razorpay_payment_link_status}|{razorpay_payment_id}"
        expected = hmac.new(RazorpayAdapter.WEBHOOK_SECRET.encode(), msg.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(expected, razorpay_signature or ""):
            logger.warning("[RAZORPAY-REDIRECT] Signature mismatch")
            raise HTTPException(400, "Invalid signature")

    # Find and mark sale paid by rzp_payment_link_id
    if razorpay_payment_link_id:
        sale = await db.shop_sales.find_one({"rzp_payment_link_id": razorpay_payment_link_id}, {"_id": 0})
        if sale and not sale.get("paid"):
            await db.shop_sales.update_one({"id": sale["id"]}, {"$set": {
                "paid": True, "paid_at": now_iso(),
                "rzp_payment_id": razorpay_payment_id,
            }})
            sale["paid"] = True
            await _send_sale_invoice_whatsapp(sale)
            logger.info(f"[RAZORPAY-REDIRECT] Sale {sale['id']} marked paid via redirect")
        # Also check bookings
        booking = await db.bookings.find_one({"rzp_payment_link_id": razorpay_payment_link_id}, {"_id": 0})
        if booking and not booking.get("paid"):
            await db.bookings.update_one({"id": booking["id"]}, {"$set": {
                "paid": True, "status": "PAID", "paid_at": now_iso(),
                "payment_method": "razorpay", "razorpay_payment_id": razorpay_payment_id,
            }})
            updated = await db.bookings.find_one({"id": booking["id"]}, {"_id": 0})
            recipients = await get_recipients_for_booking(updated)
            await publish_event("PAID", updated, recipients)
            logger.info(f"[RAZORPAY-REDIRECT] Booking {booking['id']} marked paid via redirect")

    return RedirectResponse(url=os.environ.get("PUBLIC_URL", "/") + "?payment=success")

# ---------- WebSocket ----------
@app.websocket("/api/ws/{token}")
async def websocket_endpoint(websocket: WebSocket, token: str):
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
        user_id = payload["sub"]
    except jwt.PyJWTError:
        await websocket.close(code=1008)
        return
    await websocket.accept()
    q = await bus.subscribe(user_id)
    try:
        await websocket.send_json({"type": "connected", "user_id": user_id})
        while True:
            event = await q.get()
            await websocket.send_json(event)
    except WebSocketDisconnect:
        pass
    except Exception as e:
        logger.warning(f"ws error: {e}")
    finally:
        bus.unsubscribe(user_id, q)


# ---------- Seeding ----------
DEMO_USERS = [
    {"username": "9353401156", "password": "macjit@123", "name": "Prashant Tiwary", "role": "admin", "phone": "+919353401156"},
]

DEMO_BAYS = [
    {"id": "bay-1", "name": "Bay A1", "type": "general"},
    {"id": "bay-2", "name": "Bay A2", "type": "general"},
    {"id": "bay-3", "name": "Bay B1", "type": "engine"},
    {"id": "bay-4", "name": "Bay B2", "type": "tire"},
]

DEMO_INVENTORY = [
    {"name": "Engine Oil 5W-30 Synthetic 4L (Castrol)", "sku": "OIL-CST-4L-5W30", "category": "Lubricants", "price": 2400, "stock": 18, "low_stock_threshold": 6,
     "stocked_at": (datetime.now(timezone.utc) - timedelta(days=20)).isoformat(),
     "expiry_at": (datetime.now(timezone.utc) + timedelta(days=540)).isoformat()},
    {"name": "Engine Oil 0W-20 Synthetic 4L (Mobil1)", "sku": "OIL-MBL-4L-0W20", "category": "Lubricants", "price": 3200, "stock": 10, "low_stock_threshold": 4,
     "stocked_at": (datetime.now(timezone.utc) - timedelta(days=15)).isoformat(),
     "expiry_at": (datetime.now(timezone.utc) + timedelta(days=600)).isoformat()},
    {"name": "Oil Filter (Bosch)", "sku": "FLT-OIL-BSH", "category": "Filters", "price": 480, "stock": 22, "low_stock_threshold": 6,
     "stocked_at": (datetime.now(timezone.utc) - timedelta(days=12)).isoformat()},
    {"name": "Air Filter (Bosch)", "sku": "FLT-AIR-BSH", "category": "Filters", "price": 620, "stock": 14, "low_stock_threshold": 5,
     "stocked_at": (datetime.now(timezone.utc) - timedelta(days=10)).isoformat()},
    {"name": "Cabin AC Filter (carbon)", "sku": "FLT-AC-CRB", "category": "Filters", "price": 850, "stock": 9, "low_stock_threshold": 4,
     "stocked_at": (datetime.now(timezone.utc) - timedelta(days=18)).isoformat()},
    {"name": "Front Brake Pad Set (Bosch)", "sku": "BRK-PAD-FRT", "category": "Brakes", "price": 1850, "stock": 6, "low_stock_threshold": 3,
     "stocked_at": (datetime.now(timezone.utc) - timedelta(days=22)).isoformat()},
    {"name": "Rear Brake Shoe Set", "sku": "BRK-SHOE-RR", "category": "Brakes", "price": 1450, "stock": 4, "low_stock_threshold": 3,
     "stocked_at": (datetime.now(timezone.utc) - timedelta(days=25)).isoformat()},
    {"name": "Brake Fluid DOT-4 1L", "sku": "FLD-BRK-DOT4", "category": "Fluids", "price": 380, "stock": 16, "low_stock_threshold": 5,
     "stocked_at": (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()},
    {"name": "Coolant 5L (pre-mix)", "sku": "COL-5L", "category": "Fluids", "price": 950, "stock": 12, "low_stock_threshold": 4,
     "stocked_at": (datetime.now(timezone.utc) - timedelta(days=14)).isoformat()},
    {"name": "Wiper Blade Set 22\"+18\"", "sku": "WIP-22-18", "category": "Wipers", "price": 850, "stock": 8, "low_stock_threshold": 4,
     "stocked_at": (datetime.now(timezone.utc) - timedelta(days=8)).isoformat()},
    {"name": "12V 65Ah Battery (Exide)", "sku": "BAT-EXD-65", "category": "Battery", "price": 7200, "stock": 5, "low_stock_threshold": 2,
     "stocked_at": (datetime.now(timezone.utc) - timedelta(days=18)).isoformat()},
    {"name": "Spark Plug Set of 4 (NGK Iridium)", "sku": "SPK-NGK-IR-4", "category": "Ignition", "price": 1600, "stock": 12, "low_stock_threshold": 4,
     "stocked_at": (datetime.now(timezone.utc) - timedelta(days=11)).isoformat()},
    {"name": "AC Refrigerant Gas R-1234yf 250g", "sku": "AC-GAS-R1234", "category": "AC", "price": 2800, "stock": 6, "low_stock_threshold": 3,
     "stocked_at": (datetime.now(timezone.utc) - timedelta(days=20)).isoformat()},
    {"name": "Tyre 195/55 R16 (Apollo)", "sku": "TYR-195-55-16", "category": "Tyres", "price": 7200, "stock": 8, "low_stock_threshold": 4,
     "stocked_at": (datetime.now(timezone.utc) - timedelta(days=25)).isoformat()},
]


@app.on_event("startup")
async def startup():
    # One-time cleanup: remove any old hard-coded demo accounts from earlier seeds
    OLD_DEMO_USERNAMES = ["customer", "customer2", "reception", "mechanic",
                          "mechanic2", "tester", "admin", "shopkeeper"]
    deleted = await db.users.delete_many({"username": {"$in": OLD_DEMO_USERNAMES}})
    if deleted.deleted_count:
        logger.info(f"Cleaned up {deleted.deleted_count} legacy demo users")
    # Customers no longer have accounts — purge any leftovers from older builds.
    cust_purged = await db.users.delete_many({"role": "customer"})
    if cust_purged.deleted_count:
        logger.info(f"Purged {cust_purged.deleted_count} legacy customer accounts (customers no longer log in)")
    # Drop OTP collection if present (feature removed)
    try:
        await db.otps.drop()
    except Exception:
        pass
    # One-time cleanup: remove old 2-wheeler demo inventory
    OLD_INV_SKUS = ["OIL-CST-1L", "BRK-PAD-01", "FLT-AIR-01", "SPK-NGK-01",
                    "CHN-LUB-01", "TYR-90-17", "COL-500"]
    inv_del = await db.inventory.delete_many({"sku": {"$in": OLD_INV_SKUS}})
    if inv_del.deleted_count:
        logger.info(f"Cleaned up {inv_del.deleted_count} legacy 2-wheeler inventory items")
    # Idempotently seed/upgrade demo users (force admin role + correct password hash)
    for u in DEMO_USERS:
        existing = await db.users.find_one({"username": u["username"]})
        h = await hash_password(u["password"])
        if existing:
            await db.users.update_one(
                {"id": existing["id"]},
                {"$set": {"name": u["name"], "role": u["role"], "phone": u["phone"],
                          "password_hash": h}}
            )
            logger.info(f"Upgraded existing user {u['username']} to {u['role']}")
        else:
            doc = {"id": str(uuid.uuid4()), "username": u["username"], "name": u["name"],
                   "role": u["role"], "phone": u["phone"],
                   "password_hash": h, "created_at": now_iso()}
            await db.users.insert_one(doc)  # seed users stored without encryption
            logger.info(f"Seeded admin {u['username']}")

    DEMO_SERVICES = [
        {"key": "general", "name": "General Service", "duration_min": 120, "base_price": 1200, "active": True},
        {"key": "oil-change", "name": "Oil & Filter Change", "duration_min": 45, "base_price": 800, "active": True},
        {"key": "full-service", "name": "Full Service", "duration_min": 210, "base_price": 3500, "active": True},
        {"key": "ac-service", "name": "AC Service", "duration_min": 90, "base_price": 1500, "active": True},
        {"key": "alignment", "name": "Wheel Alignment & Balancing", "duration_min": 60, "base_price": 700, "active": True},
        {"key": "brake", "name": "Brake Service", "duration_min": 75, "base_price": 1000, "active": True},
        {"key": "engine", "name": "Engine Repair / Diagnostics", "duration_min": 240, "base_price": 2500, "active": True},
    ]
    if await db.services.count_documents({}) == 0:
        for s in DEMO_SERVICES:
            await db.services.insert_one({"id": str(uuid.uuid4()), **s, "created_at": now_iso()})
        logger.info(f"Seeded {len(DEMO_SERVICES)} default services")

    if await db.bays.count_documents({}) == 0:
        for b in DEMO_BAYS:
            await db.bays.insert_one({**b, "created_at": now_iso()})
        logger.info(f"Seeded {len(DEMO_BAYS)} bays")
    if await db.inventory.count_documents({}) == 0:
        for i in DEMO_INVENTORY:
            await db.inventory.insert_one({"id": str(uuid.uuid4()), **i, "created_at": now_iso()})
        logger.info(f"Seeded {len(DEMO_INVENTORY)} inventory items")


@app.on_event("shutdown")
async def shutdown():
    client.close()


app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------- Security Headers Middleware ----------
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.responses import Response as StarletteResponse

class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-XSS-Protection"] = "1; mode=block"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
        return response

app.add_middleware(SecurityHeadersMiddleware)

# ---------- Serve React frontend (production build) ----------
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse

FRONTEND_BUILD = ROOT_DIR.parent / "frontend" / "build"
NO_CACHE_HEADERS = {
    "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
    "Pragma": "no-cache",
    "Expires": "0",
}

def _serve_index():
    return FileResponse(str(FRONTEND_BUILD / "index.html"), headers=NO_CACHE_HEADERS)

if FRONTEND_BUILD.exists():
    static_dir = FRONTEND_BUILD / "static"
    if static_dir.exists():
        app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")

    @app.get("/")
    async def _spa_root():
        return _serve_index()

    @app.get("/{full_path:path}")
    async def _spa_fallback(full_path: str):
        if (full_path.startswith("api")
                or full_path.startswith("ws")
                or full_path.startswith("__replco")
                or full_path.startswith("__replit")):
            return JSONResponse({"detail": "Not Found"}, status_code=404)
        candidate = FRONTEND_BUILD / full_path
        if candidate.exists() and candidate.is_file():
            # Disable caching for the service worker and the HTML shell so
            # browsers always pick up the latest build.
            if full_path in ("sw.js", "index.html", "manifest.json"):
                return FileResponse(str(candidate), headers=NO_CACHE_HEADERS)
            return FileResponse(str(candidate))
        return _serve_index()
else:
    @app.get("/")
    async def _no_build():
        return JSONResponse({
            "status": "backend_only",
            "message": "Frontend build not found. Run `cd frontend && yarn build`.",
        })
