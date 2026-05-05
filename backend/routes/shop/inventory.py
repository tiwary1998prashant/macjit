from server import BulkInventory, Depends, Dict, EnquiryIn, File, HTTPException, InventoryIn, List, TwilioAdapter, UploadFile, datetime, db, get_current_user, now_iso, require_roles, timezone, uuid  # noqa: F401

from fastapi import APIRouter
router = APIRouter()
# Auto-generated from routes.py section
# Section starts at line 1154

# ---------- Bulk inventory ----------
@router.post("/inventory/bulk")
async def bulk_inventory(data: BulkInventory, user=Depends(require_roles("admin", "reception"))):
    created = []
    for it in data.items:
        item = {"id": str(uuid.uuid4()), **it.model_dump(),
                "stocked_at": it.stocked_at or now_iso(),
                "created_at": now_iso()}
        await db.inventory.insert_one(dict(item))
        item.pop("_id", None)
        created.append(item)
    return {"created": len(created), "items": created}


# ---------- Bulk inventory file upload (CSV / XLSX) ----------
@router.post("/inventory/bulk-upload")
async def bulk_inventory_upload(file: UploadFile = File(...),
                                user=Depends(require_roles("admin", "reception"))):
    """Accepts .csv or .xlsx with columns:
       name, sku, category, price, stock, low_stock_threshold (optional)
       Upserts by SKU: if SKU exists, it adds to stock; otherwise creates new.
    """
    import io, csv
    filename = (file.filename or "").lower()
    raw = await file.read()
    rows: List[dict] = []
    try:
        if filename.endswith(".csv"):
            text = raw.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            rows = [r for r in reader]
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            headers = [str(h or "").strip().lower() for h in next(it)]
            for vals in it:
                if not any(v not in (None, "") for v in vals):
                    continue
                rows.append({headers[i]: (vals[i] if i < len(vals) else "") for i in range(len(headers))})
        else:
            raise HTTPException(400, "Unsupported file. Use .csv or .xlsx")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Could not parse file: {e}")

    REQUIRED = {"name", "sku", "category", "price", "stock"}
    added, updated, errors = 0, 0, []
    for idx, r in enumerate(rows, start=2):  # row 1 is header
        norm = {(k or "").strip().lower(): v for k, v in r.items()}
        missing = [k for k in REQUIRED if not str(norm.get(k, "")).strip()]
        if missing:
            errors.append({"row": idx, "error": f"Missing: {', '.join(missing)}"})
            continue
        try:
            payload = {
                "name": str(norm["name"]).strip(),
                "sku": str(norm["sku"]).strip().upper(),
                "category": str(norm["category"]).strip(),
                "price": float(norm["price"]),
                "stock": int(float(norm["stock"])),
                "low_stock_threshold": int(float(norm.get("low_stock_threshold") or 5)),
            }
        except (ValueError, TypeError) as e:
            errors.append({"row": idx, "error": f"Invalid value: {e}"})
            continue
        existing = await db.inventory.find_one({"sku": payload["sku"]}, {"_id": 0})
        if existing:
            await db.inventory.update_one(
                {"id": existing["id"]},
                {"$inc": {"stock": payload["stock"]},
                 "$set": {"name": payload["name"], "category": payload["category"],
                          "price": payload["price"],
                          "low_stock_threshold": payload["low_stock_threshold"]}})
            updated += 1
        else:
            doc = {"id": str(uuid.uuid4()), **payload,
                   "stocked_at": now_iso(), "created_at": now_iso()}
            await db.inventory.insert_one(doc)
            added += 1
    return {"added": added, "updated": updated, "errors": errors,
            "total_rows": len(rows)}


# ---------- Public Enquiries ----------
@router.post("/enquiries")
async def create_enquiry(data: EnquiryIn):
    doc = {"id": str(uuid.uuid4()), **data.model_dump(),
           "status": "new", "created_at": now_iso()}
    await db.enquiries.insert_one(dict(doc))
    doc.pop("_id", None)
    # Notify admin via SMS (if Twilio configured)
    admins = await db.users.find({"role": "admin"}, {"_id": 0}).to_list(10)
    for a in admins:
        if a.get("phone"):
            await TwilioAdapter.send_sms(
                a["phone"],
                f"MacJit enquiry from {data.name} ({data.phone}): "
                f"{(data.car_make or '')} {(data.car_model or '')} - "
                f"{(data.service_interest or 'general')}"
            )
    return doc


@router.get("/enquiries")
async def list_enquiries(user=Depends(require_roles("admin", "reception"))):
    items = await db.enquiries.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items


@router.patch("/enquiries/{enq_id}")
async def update_enquiry(enq_id: str, data: dict, user=Depends(require_roles("admin", "reception"))):
    data.pop("_id", None); data.pop("id", None)
    await db.enquiries.update_one({"id": enq_id}, {"$set": data})
    return await db.enquiries.find_one({"id": enq_id}, {"_id": 0})


# ---------- INVENTORY ----------
@router.get("/inventory")
async def list_inventory(user=Depends(get_current_user)):
    items = await db.inventory.find({}, {"_id": 0}).to_list(1000)
    items.sort(key=lambda x: x.get("stocked_at") or "")
    return items


@router.post("/inventory")
async def create_inventory(data: InventoryIn, user=Depends(require_roles("admin", "reception"))):
    item = {"id": str(uuid.uuid4()), **data.model_dump(),
            "stocked_at": data.stocked_at or now_iso(),
            "created_at": now_iso()}
    await db.inventory.insert_one(dict(item))
    item.pop("_id", None)
    return item


@router.patch("/inventory/{item_id}")
async def update_inventory(item_id: str, data: dict, user=Depends(require_roles("admin", "reception"))):
    data.pop("_id", None); data.pop("id", None)
    await db.inventory.update_one({"id": item_id}, {"$set": data})
    return await db.inventory.find_one({"id": item_id}, {"_id": 0})


@router.delete("/inventory/{item_id}")
async def delete_inventory(item_id: str, user=Depends(require_roles("admin"))):
    await db.inventory.delete_one({"id": item_id})
    return {"ok": True}


@router.get("/inventory/alerts")
async def inventory_alerts(user=Depends(require_roles("admin", "reception"))):
    items = await db.inventory.find({}, {"_id": 0}).to_list(1000)
    out_of_stock = [i for i in items if i["stock"] <= 0]
    low_stock = [i for i in items if 0 < i["stock"] <= i.get("low_stock_threshold", 5)]
    items.sort(key=lambda x: x.get("stocked_at") or "")
    by_sku: Dict[str, list] = {}
    for i in items:
        by_sku.setdefault(i["sku"], []).append(i)
    fifo = []
    for sku, group in by_sku.items():
        if len(group) > 1:
            fifo.append({"sku": sku, "use_first": group[0]["name"],
                         "stocked_at": group[0].get("stocked_at"), "id": group[0]["id"]})
    # not-sold-in-N-days alert
    threshold_days = 30
    now = datetime.now(timezone.utc)
    used_recently = set()
    async for b in db.bookings.find({"paid_at": {"$ne": None}}, {"_id": 0, "items": 1, "paid_at": 1}):
        try:
            paid = datetime.fromisoformat(b["paid_at"])
            if (now - paid).days <= threshold_days:
                for it in b.get("items", []):
                    used_recently.add(it.get("inventory_id"))
        except Exception:
            pass
    stagnant = [i for i in items if i["stock"] > 0 and i["id"] not in used_recently]
    return {"out_of_stock": out_of_stock, "low_stock": low_stock,
            "fifo": fifo, "stagnant": stagnant[:10],
            "stagnant_threshold_days": threshold_days}


